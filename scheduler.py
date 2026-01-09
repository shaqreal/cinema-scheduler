"""
Weekly Cinema Scheduler (v1.0)

This script updates a *single* master schedule workbook (e.g., "Current Schedule.xlsx")
using a bookings-export workbook (e.g., "Bookings Export.xlsx").

What it does (high level)
-------------------------
For each theater (Circuit + Theatre Name + City + ST), the script:

1) Determines:
   - Current play week: max(Playwk)
   - Previous play week: the latest Playwk < current

2) Builds schedules:
   - Previous week rows: writes FSS values (and preserves the titles layout already on the sheet)
     for Hold + Final rows.
   - Current week rows: writes only titles (no FSS), using a packing algorithm that stacks
     titles within each screen cell. NEW titles are highlighted.

3) Applies consistent styling:
   - Header styling (site cell and screen headers)
   - Week banding (alternating blues by ISO week parity)
   - Borders and wrapped alignment

4) Handles an "Alternative Content" column (if present) by writing event rows for both weeks.

CLI Options
-----------
--bookings-sheet NAME          Sheet name in bookings workbook (default: "Bookings")
--theater-prefix-regex REGEX    Optional regex to strip prefixes from theater names
--sheet-name-format FORMAT     Sheet name format string (default: "{theatre}, {state}")
--include-city-in-sheet-name   Include city in sheet names to avoid collisions

Column Aliases
--------------
The script supports flexible column naming via aliases. Common aliases include:
- "Theatre Name" / "Theater Name" / "Location" / "Site" / "Venue"
- "ST" / "State" / "Province"
- "Playwk" / "Play Week" / "Week Start" / "WeekOf"
- "WK#" / "WK" / "Week#" / "Week #" / "WeekNum"
- "FSS" / "FriSatSun" / "WeekendGross" / "Weekend"
- "Status" / "Booking Status" / "Run Status"
- "Circuit" / "Chain" / "Company" / "Exhibitor"
- "DIST" / "Distributor" / "Dist."

Usage
-----
python cinema_scheduler.py "Bookings Export.xlsx" "Current Schedule.xlsx"
  python cinema_scheduler.py "Bookings.xlsx" "Schedule.xlsx" --bookings-sheet "Data"

Expected input columns (canonical names)
----------------------------------------
Required:
- Circuit (or alias: Chain, Company, Exhibitor)
- Theatre Name (or alias: Theater Name, Location, Site, Venue)
- City
- ST (or alias: State, Province)
- Title
- Status (or alias: Booking Status, Run Status)
- Playwk (or alias: Play Week, Week Start, WeekOf)
- WK# (or alias: WK, Week#, Week #, WeekNum)
- FSS (or alias: FriSatSun, WeekendGross, Weekend)

Optional:
- Total
- DIST (or alias: Distributor, Dist.)
- Standard, ATMOS, IMAX, etc. (any screen-unit columns)

Notes / caveats
---------------
- Rich text in Excel: openpyxl does NOT reliably round-trip per-run formatting across saves.
  To preserve per-line "(Final)" red formatting, this script re-applies rich text formatting
  to all historical FSS cells after updates.
- Excel safety: strings are sanitized to remove illegal XML/control characters that can
  prevent Excel from opening a workbook.

Dependencies
------------
- pandas
- openpyxl

Author
------
Sasha (sasha2820@outlook.com)

License
-------
MIT License. See LICENSE for details.
"""

from __future__ import annotations

import argparse
import math
import os
import re
from datetime import datetime, date
from typing import Dict, Iterable, List, Optional, Sequence, Tuple, Union

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ---------------------------------------------------------------------------
# Configuration  
# ---------------------------------------------------------------------------

BOOKINGS_SHEET_NAME = "Bookings"

# Metadata cell for theater identification (stable across runs)
META_CELL = "AA1"
META_COL = "AA"


# --- Template-based default screen counts (minimum, never a max) ---
# If a template workbook exists alongside the script (or in CWD), we use it to seed
# each theater's *minimum* screen headers (e.g., Rotunda=7, Frederick=10, Leitersburg=10).
# We still expand beyond this if the computed layout requires more screens.
TEMPLATE_WORKBOOK_CANDIDATES = [
    "Warehouse Cinemas - OG.xlsx",
    "Warehouse Cinemas - Current.xlsx",
]

_TEMPLATE_WB_CACHE = None
_TEMPLATE_SCREENS_CACHE = {}

def _find_template_workbook_path() -> Optional[str]:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = []
    for fname in TEMPLATE_WORKBOOK_CANDIDATES:
        candidates.append(os.path.join(script_dir, fname))
        candidates.append(os.path.join(os.getcwd(), fname))
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

def _load_template_wb() -> Optional[Workbook]:
    global _TEMPLATE_WB_CACHE
    if _TEMPLATE_WB_CACHE is not None:
        return _TEMPLATE_WB_CACHE
    path = _find_template_workbook_path()
    if not path:
        _TEMPLATE_WB_CACHE = None
        return None
    try:
        _TEMPLATE_WB_CACHE = load_workbook(path)
        return _TEMPLATE_WB_CACHE
    except Exception:
        _TEMPLATE_WB_CACHE = None
        return None

def get_min_screen_count_from_template(theater_name: Optional[str], state: Optional[str]) -> int:
    """
    Returns a *minimum* screen count for the theater based on the OG/Current template, if found.
    Never caps: real required screens can exceed this.

    We match template sheets like "Rotunda, MD" by using the short theatre name:
      "Warehouse Cinemas Rotunda" -> "Rotunda, MD"
    """
    if not theater_name or not state:
        return 1

    short = re.sub(r"^Warehouse\s+Cinemas\s+", "", str(theater_name).strip(), flags=re.IGNORECASE).strip()
    if not short:
        return 1

    key = f"{short.lower()}|{str(state).strip().lower()}"
    if key in _TEMPLATE_SCREENS_CACHE:
        return _TEMPLATE_SCREENS_CACHE[key]

    wb = _load_template_wb()
    if wb is None:
        _TEMPLATE_SCREENS_CACHE[key] = 1
        return 1

    # Try exact match first: "{short}, {state}"
    target = f"{short}, {str(state).strip()}"
    sheet = None
    if target in wb.sheetnames:
        sheet = wb[target]
    else:
        # Fallback: any sheet whose normalized key matches
        tkey = normalize_sheet_key(target)
        for sn in wb.sheetnames:
            if normalize_sheet_key(sn) == tkey:
                sheet = wb[sn]
                break

    if sheet is None:
        _TEMPLATE_SCREENS_CACHE[key] = 1
        return 1

    # Count integer screen headers on row 2 (col>=2)
    count = 0
    for col in range(2, sheet.max_column + 1):
        v = sheet.cell(row=2, column=col).value
        if v is None:
            continue
        s = str(v).strip()
        if re.fullmatch(r"\d+", s):
            count += 1

    _TEMPLATE_SCREENS_CACHE[key] = max(1, count)
    return _TEMPLATE_SCREENS_CACHE[key]

YELLOW_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FINAL_FONT = Font(color="FFFF0000", bold=True)

# Week banding fills (alternating shades)
WEEK_BLUE_A = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
WEEK_BLUE_B = PatternFill(start_color="FFE7F3FF", end_color="FFE7F3FF", fill_type="solid")

# Header and city styling
HEADER_BLUE = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
CITY_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # same as YELLOW_FILL
NEW_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
NEW_ORANGE = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")

# Border styling
THIN_SIDE = Side(style="thin", color="000000")
GRID_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# Alignment for wrapped text
WRAP_TOP = Alignment(wrap_text=True, horizontal="center", vertical="center")

# Rich text inline fonts
INLINE_NORMAL = InlineFont()
INLINE_FINAL_RED = InlineFont(color="FFFF0000", b=True)

# Screen unit columns (auto-detected from bookings DataFrame)
SCREEN_UNIT_COLS: List[str] = []


# The first detected unit column is treated as the "standard/digital" room pool for that export.
# Remaining detected unit columns are treated as separate premium room pools, appended in order.
STANDARD_UNIT_COL: str = ""
PREMIUM_UNIT_COLS: List[str] = []
# Columns used throughout the script.
COL_TITLE = "Title"
COL_STATUS = "Status"
COL_PLAYWK = "Playwk"
COL_WKNUM = "WK#"
COL_STANDARD = "Standard"
COL_ATMOS = "ATMOS"
COL_TOTAL = "Total"
COL_FSS = "FSS"
COL_DIST = "DIST"


# ---------------------------------------------------------------------------
# Styling Helpers
# ---------------------------------------------------------------------------

def week_band_fill(play_week: datetime) -> PatternFill:
    """
    Pick week band color deterministically (stable across runs).
    
    Uses ISO week number parity for stable alternating shading.
    """
    iso_week = play_week.isocalendar().week
    return WEEK_BLUE_A if (iso_week % 2 == 0) else WEEK_BLUE_B


def apply_week_band(
    ws: Worksheet,
    titles_row: int,
    fss_row: int,
    screen_cols: List[int],
    alt_content_col: Optional[int],
    unplayed_col: Optional[int],
    play_week: datetime,
) -> None:
    """
    Apply week band fill to a week block (titles row + FSS row).
    
    Only fills screen columns (and Unplayed/Alt Content if present).
    Does NOT fill column A (date column).
    Does NOT overwrite NEW fills (yellow/orange) or CITY_YELLOW.
    """
    fill = week_band_fill(play_week)
    cols = list(screen_cols)
    if alt_content_col is not None:
        cols.append(alt_content_col)
    if unplayed_col is not None:
        cols.append(unplayed_col)

    for c in cols:
        # Titles row: only apply week band if blank or already week band fill
        titles_cell = ws.cell(row=titles_row, column=c)
        if is_blank_fill(titles_cell.fill) or is_week_band_fill(titles_cell.fill):
            titles_cell.fill = fill
        # Otherwise preserve existing fill (NEW yellow/orange etc)
        
        # FSS row: same rule (blank or week band) before applying
        fss_cell = ws.cell(row=fss_row, column=c)
        if is_blank_fill(fss_cell.fill) or is_week_band_fill(fss_cell.fill):
            fss_cell.fill = fill
        # Otherwise preserve existing fill


def apply_borders_to_week_block(
    ws: Worksheet,
    titles_row: int,
    fss_row: int,
    cols: List[int],
) -> None:
    """Apply borders to all cells in a week block (titles row + FSS row)."""
    for r in (titles_row, fss_row):
        for c in cols:
            ws.cell(row=r, column=c).border = GRID_BORDER
            ws.cell(row=r, column=c).alignment = WRAP_TOP


def apply_border_to_date_col(ws: Worksheet, titles_row: int, fss_row: int) -> None:
    """Apply borders + centered alignment to column A date cells for a week block."""
    for r in (titles_row, fss_row):
        cell = ws.cell(row=r, column=1)
        cell.border = GRID_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
def style_headers(
    ws: Worksheet,
    screen_cols: List[int],
    alt_col: Optional[int],
    unplayed_col: Optional[int],
) -> None:
    """
    Style header row (A1 and row 2) with fills, fonts, and borders.
    
    A1: CITY_YELLOW fill, bold, border
    Row 2: HEADER_BLUE fill, bold, centered, border for col A and each screen col (+ alt/unplayed if present)
    """
    # Style A1 (city/theater name)
    a1_cell = ws.cell(row=1, column=1)
    a1_cell.fill = CITY_YELLOW
    a1_cell.font = Font(bold=True)
    a1_cell.border = GRID_BORDER
    
    # Style row 2 headers
    cols_to_style = [1] + screen_cols  # Column A + screen columns
    if alt_col is not None:
        cols_to_style.append(alt_col)
    if unplayed_col is not None:
        cols_to_style.append(unplayed_col)
    
    for col in cols_to_style:
        header_cell = ws.cell(row=2, column=col)
        header_cell.fill = HEADER_BLUE
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = GRID_BORDER


def _fill_rgb(fill: PatternFill) -> str:
    """Get RGB hex string from a PatternFill."""
    if fill is None or fill.start_color is None:
        return ""
    return fill.start_color.rgb or ""


def _is_default_fill(fill: PatternFill) -> bool:
    """Check if fill is default/empty."""
    if fill is None:
        return True
    rgb = _fill_rgb(fill)
    return not rgb or rgb == "00000000"  # Default/transparent


def is_blank_fill(fill: PatternFill) -> bool:
    """Check if fill is blank/default (no patternType or no rgb / rgb == "00000000")."""
    if fill is None:
        return True
    if not hasattr(fill, 'fill_type') or fill.fill_type is None:
        return True
    rgb = _fill_rgb(fill)
    return not rgb or rgb == "00000000"


def is_week_band_fill(fill: PatternFill) -> bool:
    """Check if fill RGB matches WEEK_BLUE_A or WEEK_BLUE_B."""
    if fill is None:
        return False
    rgb = _fill_rgb(fill)
    if not rgb:
        return False
    week_blue_a_rgb = _fill_rgb(WEEK_BLUE_A)
    week_blue_b_rgb = _fill_rgb(WEEK_BLUE_B)
    return rgb == week_blue_a_rgb or rgb == week_blue_b_rgb


def restyle_final_fss_cells(ws: Worksheet, cols: list[int], start_row: int = 3) -> None:
    """
    Re-apply rich text formatting to FSS cells containing "(Final)" across all old weeks.
    
    This ensures per-line red formatting persists across multiple runs, since openpyxl
    does not reliably round-trip rich text.
    """
    for r in range(start_row, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, (datetime, date)):  # titles row
            fss_row = r + 1
            if fss_row > ws.max_row:
                continue
            for c in cols:
                cell = ws.cell(row=fss_row, column=c)
                v = cell.value
                if isinstance(v, str) and "(final)" in v.lower():
                    # Sanitize text before parsing
                    v = sanitize_excel_text(v)
                    # Split into blocks on blank lines (double newlines)
                    blocks = [b.strip() for b in re.split(r"\n\s*\n", v) if b.strip()]
                    set_rich_text_multiline(cell, blocks)
                    cell.alignment = WRAP_TOP
                    cell.border = GRID_BORDER


def sanitize_excel_text(s: object) -> str:
    """
    Sanitize text to be Excel-safe by removing illegal characters and normalizing.
    
    This is the single source of truth for text sanitization. All strings written
    to Excel cells should pass through this function.
    
    Args:
        s: Input value (can be None, string, or other)
    
    Returns:
        Sanitized string safe for Excel
    """
    if s is None:
        return ""
    txt = str(s)
    
    # Remove illegal XML/control chars Excel can't load (from openpyxl)
    txt = ILLEGAL_CHARACTERS_RE.sub("", txt)
    
    # Remove additional ASCII control chars: 0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F
    # (0x09=tab, 0x0A=LF, 0x0D=CR are allowed and will be normalized below)
    import string
    control_chars = ''.join(chr(i) for i in range(32) if i not in (9, 10, 13))
    for char in control_chars:
        txt = txt.replace(char, "")
    
    # Normalize line endings: \r\n and \r -> \n
    txt = txt.replace("\r\n", "\n").replace("\r", "\n")
    
    # Remove zero-width chars that sometimes sneak in
    txt = txt.replace("\u200b", "")  # Zero-width space
    txt = txt.replace("\u200c", "")  # Zero-width non-joiner
    txt = txt.replace("\u200d", "")  # Zero-width joiner
    txt = txt.replace("\ufeff", "")  # Zero-width no-break space (BOM)
    
    # Strip quotes that wrap the whole cell if it contains newlines
    if txt.startswith('"') and txt.endswith('"') and "\n" in txt:
        txt = txt[1:-1]
    
    return txt


def make_safe_sheet_title(raw: str, existing: set[str]) -> str:
    """
    Create a safe Excel sheet title that meets all Excel restrictions.
    
    Excel sheet name restrictions:
    - Cannot contain: [ ] : * ? / \
    - Max 31 characters
    - Cannot be empty
    - Must be unique within the workbook
    
    Args:
        raw: Raw sheet title to make safe
        existing: Set of existing sheet names in the workbook (for uniqueness)
    
    Returns:
        Safe sheet title that is unique and valid
    """
    if not raw:
        raw = "Sheet"
    
    # Remove illegal Excel sheet name characters: []:*?/\
    safe = re.sub(r'[\[\]:*?/\\]', '_', raw)
    
    # Strip leading/trailing whitespace
    safe = safe.strip()
    
    # If empty after cleaning, use default
    if not safe:
        safe = "Sheet"
    
    # Truncate to 31 characters max
    if len(safe) > 31:
        safe = safe[:31]

    # Remove trailing commas/spaces created by truncation
    safe = safe.rstrip(' ,')
    
    # Ensure uniqueness by appending suffixes like " (2)", " (3)" while staying <= 31 chars
    base = safe
    counter = 2
    while safe in existing:
        # Calculate suffix length: " (N)" where N can be multi-digit
        suffix = f" ({counter})"
        # Truncate base if needed to leave room for suffix
        max_base_len = 31 - len(suffix)
        if max_base_len <= 0:
            # If suffix itself is too long, just use counter
            safe = f"Sheet{counter}"
        else:
            truncated_base = base[:max_base_len]
            safe = truncated_base + suffix
            # If still too long, truncate more aggressively
            if len(safe) > 31:
                safe = safe[:31]
            safe = safe.rstrip(' ,')
        counter += 1
        # Safety limit to prevent infinite loop
        if counter > 1000:
            safe = f"Sheet{hash(base) % 10000}"
            break
    
    return safe


def sanitize_workbook_inplace(wb: Workbook) -> None:
    """
    Sanitize all string values in a workbook to remove illegal Excel characters.
    
    This function scrubs the entire workbook, including:
    - All cell string values
    - Rich text cell values (sanitizes each TextBlock)
    - Sheet titles (only if they contain illegal chars)
    
    Does NOT modify formulas (cells starting with "=").
    
    Args:
        wb: Workbook to sanitize in-place
    """
    for ws in wb.worksheets:
        # DO NOT rename sheet titles - this causes duplicates on next run
        # Sheet titles are only sanitized when creating NEW sheets
        
        # Sanitize all cell values
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                # Skip formulas
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                
                # Handle string values
                if isinstance(cell.value, str):
                    sanitized = sanitize_excel_text(cell.value)
                    if sanitized != cell.value:
                        cell.value = sanitized
                
                # Handle rich text
                elif isinstance(cell.value, CellRichText):
                    # Rebuild rich text with sanitized blocks
                    new_rt = CellRichText()
                    for block in cell.value:
                        if isinstance(block, TextBlock):
                            sanitized_text = sanitize_excel_text(block.text)
                            new_rt.append(TextBlock(block.font, sanitized_text))
                        else:
                            # Fallback for other block types
                            sanitized_text = sanitize_excel_text(str(block))
                            new_rt.append(TextBlock(INLINE_NORMAL, sanitized_text))
                    cell.value = new_rt


def validate_workbook(wb: Workbook, max_reports: int = 10) -> List[Tuple[str, str, str]]:
    """
    Scan workbook for any remaining illegal characters and report them.
    
    Args:
        wb: Workbook to validate
        max_reports: Maximum number of offending cells to report
    
    Returns:
        List of tuples (sheet_name, cell_coordinate, preview_text) for offending cells
    """
    issues = []
    control_chars_pattern = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\u200b\u200c\u200d\ufeff]')
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                # Skip formulas
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                
                # Check string values
                if isinstance(cell.value, str):
                    if control_chars_pattern.search(cell.value) or ILLEGAL_CHARACTERS_RE.search(cell.value):
                        preview = cell.value[:50] + ("..." if len(cell.value) > 50 else "")
                        issues.append((ws.title, cell.coordinate, preview))
                        if len(issues) >= max_reports:
                            return issues
                
                # Check rich text
                elif isinstance(cell.value, CellRichText):
                    for block in cell.value:
                        text_to_check = block.text if isinstance(block, TextBlock) else str(block)
                        if control_chars_pattern.search(text_to_check) or ILLEGAL_CHARACTERS_RE.search(text_to_check):
                            preview = text_to_check[:50] + ("..." if len(text_to_check) > 50 else "")
                            issues.append((ws.title, cell.coordinate, preview))
                            if len(issues) >= max_reports:
                                return issues
    
    return issues


def set_rich_text_multiline(cell, lines: list[str]) -> None:
    """
    Set a cell value to rich text with multiple lines.
    Lines containing "(Final)" are rendered in red.
    Uses double newlines between entries for proper spacing.
    
    IMPORTANT: Newlines are included within TextBlock text (not as separate whitespace-only runs)
    to ensure Excel preserves spacing with xml:space="preserve".
    """
    if not lines:
        cell.value = None
        return
    
    rt = CellRichText()
    for i, line in enumerate(lines):
        # Sanitize each line
        line = sanitize_excel_text(line)
        if not line.strip():
            continue
        
        # Determine if this line contains "(Final)"
        is_final = "(final)" in line.lower()
        font = INLINE_FINAL_RED if is_final else INLINE_NORMAL
        
        # Include newlines within the TextBlock text (not as separate runs)
        # Add double newline separator before all lines except the first
        if i > 0:
            text = "\n\n" + line
        else:
            text = line
        
        rt.append(TextBlock(font, text))
    
    cell.value = rt


def resolve_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Resolve column aliases to canonical names.
    
    Args:
        df: DataFrame with potentially aliased column names
    
    Returns:
        DataFrame with columns renamed to canonical names
    
    Raises:
        ValueError: If required columns are missing after resolution
    """
    # Trim whitespace around column headers
    df.columns = df.columns.str.strip()
    
    # Create a mapping from aliases to canonical names (case-insensitive)
    alias_to_canonical: Dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_to_canonical[alias.lower()] = canonical
    
    # Build rename mapping
    rename_map: Dict[str, str] = {}
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in alias_to_canonical:
            canonical = alias_to_canonical[col_lower]
            if col != canonical:
                rename_map[col] = canonical
    
    # Rename columns
    if rename_map:
        df = df.rename(columns=rename_map)
    
    # Check for required columns
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(
            f"Missing required columns after alias resolution: {missing}. "
            f"Found columns: {list(df.columns)}"
        )
    
    return df


def load_bookings(path: str, sheet_name: str = BOOKINGS_SHEET_NAME) -> pd.DataFrame:
    """
    Load the workbook into a DataFrame and normalize columns.
    
    Filters out:
    - Rows where Circuit is NaN
    - Rows where Circuit (case-insensitive) is "circuit" (repeated header rows)
    - Rows where Circuit (case-insensitive) is "total" (separator rows)
    """
    df = pd.read_excel(path, sheet_name=sheet_name)
    
    # Resolve column aliases to canonical names
    df = resolve_columns(df)
    
    # Filter out header rows and total rows
    if "Circuit" in df.columns:
        # Drop rows where Circuit is NaN
        df = df[df["Circuit"].notna()].copy()
        
        # Convert Circuit to string for case-insensitive comparison (after filtering NaN)
        circuit_str = df["Circuit"].astype(str).str.strip().str.lower()
        
        # Drop rows where Circuit is "circuit" (repeated headers) or "total" (separator rows)
        # Use a single mask to avoid reindexing warning
        df = df[(circuit_str != "circuit") & (circuit_str != "total")].copy()
    
    # Normalize identity fields: Circuit, Theatre Name, City, ST
    # Strip, collapse whitespace, strip trailing commas/spaces
    if "Circuit" in df.columns:
        df["Circuit"] = df["Circuit"].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True).str.rstrip(', ')
    if "Theatre Name" in df.columns:
        df["Theatre Name"] = df["Theatre Name"].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True).str.rstrip(', ')
    if "City" in df.columns:
        df["City"] = df["City"].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True).str.rstrip(', ')
    if "ST" in df.columns:
        # ST: strip, uppercase, remove non-letters
        df["ST"] = df["ST"].astype(str).str.strip().str.upper().str.replace(r'[^A-Z]', '', regex=True)
    
    # Parse Playwk and drop NaT rows
    df[COL_PLAYWK] = pd.to_datetime(df[COL_PLAYWK], errors="coerce")
    df = df[df[COL_PLAYWK].notna()].copy()
    
    # Coerce numeric columns
    # Coerce core numeric columns (unit columns are detected later and then coerced).
    numeric_cols = [COL_TOTAL, COL_WKNUM, COL_FSS, "Fri", "Sat", "Sun"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")# IMPORTANT:
    # Do NOT merge ATMOS into Standard.
    # We need ATMOS units to remain separate so we can append ATMOS screens
    # after Standard screens in the layout.
    
    # Filter out notes/catalog rows: WK# > 100 (like "Lost Boys, The" with WK#=1995)
    if COL_WKNUM in df.columns:
        def is_catalog_row(wk_val):
            if pd.isna(wk_val):
                return False
            try:
                wk_int = int(wk_val)
                return wk_int > 100
            except (TypeError, ValueError):
                return False
        
        catalog_mask = df[COL_WKNUM].apply(is_catalog_row)
        df = df[~catalog_mask].copy()
    
    return df


def is_event_status(status: Optional[str]) -> bool:
    """
    Return True if the status denotes an event row.
    
    Event-type rows are:
    - Status contains "event" (case-insensitive), OR
    - Status does not start with "new" and does not contain "hold" or "final"
      (catches weird one-off status strings like dates)
    """
    if not isinstance(status, str):
        return False
    status_lower = status.lower()
    if "event" in status_lower:
        return True
    # Check for weird statuses that aren't new/hold/final
    if not status_lower.startswith("new") and "hold" not in status_lower and "final" not in status_lower:
        return True
    return False


def is_event_row(row: pd.Series) -> bool:
    """Row-level event detector.

    The exports sometimes use non-standard Status strings (date ranges, show requirements)
    for regular rows that are NOT true "events", but those rows also typically have *zero*
    screen units. If we classify all non-(new/hold/final) statuses as events,
    we end up dropping real movie rows and then incorrectly pulling in prior-week holdovers.

    Rules:
      1) If Status contains "event" -> event.
      2) Otherwise, if the row has *zero* units (across all detected unit columns) and
         the FSS is empty/zero -> treat as event-ish (alt content / keys-needed / placeholders).
    """
    try:
        status = normalize_status(row.get(COL_STATUS))
    except Exception:
        status = ""
    if "event" in status:
        return True

    # If it doesn't consume any screen units, treat it as non-screen content.
    try:
        units = float(compute_units(row))
    except Exception:
        units = 0.0
    if units <= 1e-6:
        fss = row.get(COL_FSS)
        try:
            fss_val = float(fss) if pd.notna(fss) else 0.0
        except Exception:
            fss_val = 0.0
        if fss_val <= 1e-6:
            return True

    return False


def normalize_status(status: Optional[str]) -> str:
    """Return a lowercase status string for consistent comparisons."""
    return (status or "").strip().lower()


def detect_screen_unit_cols(df: pd.DataFrame) -> List[str]:
    """
    Infer screen-unit columns in *file order* by reading the columns that appear
    **after the Comments column**.

    This is intentionally circuit-agnostic:
      - Warehouse exports: Comments -> Standard, ATMOS
      - Maya exports: Comments -> Digital, ATMOS, MPX, DBOX, 3D, ...
      - Other exports: whatever unit columns exist after Comments

    Rules:
      - Take columns strictly after Comments (or Comment/Notes as fallback).
      - Keep their left-to-right order.
      - Keep only columns that contain ANY positive numeric value.
      - Exclude obvious non-unit fields (Total, FSS, Fri/Sat/Sun, WK#, etc.).

    If no Comments-like column exists, fall back to scanning all columns in order
    with the same numeric/positive/exclusion rules.
    """
    if df is None or df.empty:
        return []

    # Identify the "Comments" column (case-insensitive), using a few common variants.
    comments_idx = None
    comment_names = {"comments", "comment", "notes", "note"}
    for i, c in enumerate(df.columns):
        if str(c).strip().lower() in comment_names:
            comments_idx = i
            break

    # Columns we never want to treat as unit columns
    exclude = {
        "total", "fss", "fri", "sat", "sun", "wk#", "wk", "week", "playwk",
        "title", "status", "dist", "circuit", "theatre", "theater",
        "theatre name", "theater name", "circuit name", "comments", "comment", "notes", "note"
    }

    def is_unit_col(col_name: str) -> bool:
        if col_name is None:
            return False
        cn = str(col_name).strip()
        if not cn:
            return False
        l = cn.lower()
        if l in exclude:
            return False
        if l.startswith("unnamed"):
            return False
        # If the column is completely non-numeric, ignore it
        numeric_series = pd.to_numeric(df[cn], errors="coerce")
        return (numeric_series > 0).any()

    # Preferred: read strictly after Comments
    candidates: List[str]
    if comments_idx is not None and comments_idx + 1 < len(df.columns):
        candidates = [str(c) for c in df.columns[comments_idx + 1:]]
    else:
        # Fallback: scan everything in order
        candidates = [str(c) for c in df.columns]

    unit_cols: List[str] = []
    for c in candidates:
        try:
            if is_unit_col(c):
                unit_cols.append(c)
        except Exception:
            continue

    return unit_cols

def compute_units(row: pd.Series) -> float:
    """
    Compute total screen units for a row.

    IMPORTANT BEHAVIOR (requested):
    - Prefer summing the detected screen-unit columns (Standard, ATMOS, etc.).
    - Do NOT use "Total" as an override when format columns exist, because "Total"
      collapses Standard + ATMOS and breaks proper separation.
    - If no unit columns were detected at all, fall back to Total as a last resort.
    """
    # Sum detected unit columns if any exist
    if SCREEN_UNIT_COLS:
        units = 0.0
        found_any = False
        for col in SCREEN_UNIT_COLS:
            if col in row.index:
                value = row.get(col)
                if pd.notna(value):
                    found_any = True
                    try:
                        units += float(value)
                    except (TypeError, ValueError):
                        # Non-numeric cell, ignore
                        pass
        if found_any:
            return units

    # Last-resort fallback if no unit columns exist in the dataset
    total = row.get(COL_TOTAL)
    if pd.notna(total):
        try:
            return float(total)
        except (TypeError, ValueError):
            return 0.0
    return 0.0



def explode_row_into_pieces(row: pd.Series) -> List[pd.Series]:
    """
    Explode a booking row into format-specific pieces.
    
    For each unit column in SCREEN_UNIT_COLS (ex: Standard, ATMOS):
    - if value > 0: create a piece with __units = value and __format = column name
    - Keep all other fields (Title, Status, WK#, FSS, DIST, etc.)
    
    If nothing positive in unit cols:
    - if Total > 0: create one piece as Standard with __units=Total and __format="Standard"
    
    Args:
        row: Booking row to explode
    
    Returns:
        List of piece rows (Series objects) with __units and __format fields
    """
    pieces = []
    
    # Check each unit column
    has_format_units = False
    for col in SCREEN_UNIT_COLS:
        if col in row:
            value = row.get(col)
            if pd.notna(value) and float(value) > 0:
                has_format_units = True
                # Create a piece for this format
                piece = row.copy()
                piece["__units"] = float(value)
                piece["__format"] = col  # e.g., "Standard", "ATMOS", "IMAX"
                pieces.append(piece)
    
    # If no format units found, fall back to Total as Standard
    if not has_format_units:
        total = row.get(COL_TOTAL)
        if pd.notna(total) and float(total) > 0:
            piece = row.copy()
            piece["__units"] = float(total)
            piece["__format"] = "Standard"  # Default format
            pieces.append(piece)
    
    return pieces


def compute_format_units(rows_df: pd.DataFrame) -> Dict[str, float]:
    """
    Compute total units by format from booking rows.
    
    For each row, explodes into format-specific pieces and sums units by format.
    
    Args:
        rows_df: DataFrame of booking rows
    
    Returns:
        Dict mapping format name -> total units (e.g., {"Standard": 9.0, "ATMOS": 1.0})
    """
    fmt_units: Dict[str, float] = {}
    
    for _, row in rows_df.iterrows():
        pieces = explode_row_into_pieces(row)
        for piece in pieces:
            fmt = piece.get("__format", "Standard")
            units = float(piece.get("__units", 0.0))
            fmt_units[fmt] = fmt_units.get(fmt, 0.0) + units
    
    return fmt_units


def apply_format_labels_to_headers(
    ws: Worksheet, 
    screen_cols: List[int], 
    fmt_units: Dict[str, float], 
    header_row: int = 2
) -> None:
    """
    Apply format labels to screen headers based on units needed.
    
    Labels the rightmost available screen columns with premium format labels
    (e.g., "10 Atmos", "7 IMAX") based on units required for each format.
    
    Args:
        ws: Worksheet to update
        screen_cols: List of screen column indices (1-based)
        fmt_units: Dict mapping format name -> total units needed
        header_row: Row number containing headers (default 2)
    """
    if not screen_cols or not fmt_units:
        return
    
    # Premium formats in priority order (excluding Standard and ATMOS)
    # NOTE: ATMOS is not treated as a separate format - it's merged into Standard
    PREMIUM_FORMATS = ["IMAX", "RPX", "4DX", "DBOX", "SCREENX", "DOLBY", "LASER", "PLF"]
    
    # Regex patterns for extracting screen numbers
    pure_int_pattern = re.compile(r'^\s*\d+\s*$')
    starts_with_int_pattern = re.compile(r'^\s*(\d+)\b')
    
    # Build list of available screen columns (those without premium labels already)
    available_cols = []
    for col in screen_cols:
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is None:
            # Blank - available for labeling
            available_cols.append((col, None))
        else:
            cell_str = str(cell_value).strip()
            # Check if it's a pure integer (can be labeled)
            if pure_int_pattern.match(cell_str):
                try:
                    screen_num = int(cell_str)
                    available_cols.append((col, screen_num))
                except (ValueError, TypeError):
                    available_cols.append((col, None))
            else:
                # Check if it already has a premium label
                cell_lower = cell_str.lower()
                has_premium_label = False
                for fmt in PREMIUM_FORMATS:
                    if fmt.lower() in cell_lower:
                        has_premium_label = True
                        break
                # If it doesn't have a premium label, it's available
                if not has_premium_label:
                    # Try to extract screen number
                    match = starts_with_int_pattern.match(cell_str)
                    if match:
                        try:
                            screen_num = int(match.group(1))
                            available_cols.append((col, screen_num))
                        except (ValueError, TypeError):
                            available_cols.append((col, None))
                    else:
                        available_cols.append((col, None))
    
    # Process premium formats in priority order, assigning rightmost columns
    for fmt in PREMIUM_FORMATS:
        if fmt not in fmt_units or fmt_units[fmt] <= 0:
            continue
        
        # Calculate screens needed for this format
        needed = max(1, int(math.ceil(fmt_units[fmt])))
        
        # Take rightmost available columns
        if len(available_cols) < needed:
            # Not enough columns, skip this format
            continue
        
        # Get rightmost columns
        cols_to_label = available_cols[-needed:]
        available_cols = available_cols[:-needed]
        
        # Format label: title-case the format name (e.g., "Atmos", "Imax")
        fmt_label = fmt.title() if fmt != "IMAX" else "IMAX"  # Keep IMAX uppercase
        if fmt == "4DX":
            fmt_label = "4DX"  # Keep 4DX as-is
        
        # Label each column
        for col, screen_num in cols_to_label:
            cell = ws.cell(row=header_row, column=col)
            
            # Determine screen number to use
            if screen_num is not None:
                screen_num_to_use = screen_num
            else:
                # Use 1-based position in screen_cols
                try:
                    screen_num_to_use = screen_cols.index(col) + 1
                except ValueError:
                    screen_num_to_use = col - 1  # Approximate from column index
            
            # Create label: "<screen_number> <format>" (e.g., "10 Atmos")
            new_label = f"{screen_num_to_use} {fmt_label}"
            
            # Update cell value (preserve styling)
            cell.value = sanitize_excel_text(new_label)


def get_screen_pools(ws: Worksheet, screen_cols: List[int], header_row: int = 2) -> Dict[str, List[int]]:
    """
    Detect which screen columns belong to which format pool.
    
    From worksheet header text:
    - Pure integer headers => pool "Standard"
    - Headers containing "atmos" (case-insensitive) => pool "ATMOS"
    - Headers containing "imax" (case-insensitive) => pool "IMAX"
    - Headers containing "rpx" (case-insensitive) => pool "RPX"
    - etc. (matches unit column names)
    
    Args:
        ws: Worksheet to analyze
        screen_cols: List of screen column indices (1-based)
        header_row: Row number containing headers (default 2)
    
    Returns:
        Dict mapping format_key -> list of column indices
        format_key matches unit column names (e.g., "ATMOS", "Standard")
    """
    pools: Dict[str, List[int]] = {}
    
    if header_row > ws.max_row:
        # No headers, assume all are Standard
        pools["Standard"] = screen_cols
        return pools
    
    for col in screen_cols:
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is None:
            # Default to Standard for empty headers
            if "Standard" not in pools:
                pools["Standard"] = []
            pools["Standard"].append(col)
            continue
        
        cell_str = str(cell_value).strip().lower()
        
        # Check if it's a pure integer (Standard screen)
        try:
            int(cell_str)
            # Pure integer = Standard
            if "Standard" not in pools:
                pools["Standard"] = []
            pools["Standard"].append(col)
            continue
        except (ValueError, TypeError):
            pass
        
        # Check for format keywords in header
        format_found = None
        for unit_col in SCREEN_UNIT_COLS:
            unit_col_lower = unit_col.lower()
            if unit_col_lower in cell_str:
                format_found = unit_col  # Use original case (e.g., "ATMOS")
                break
        
        if format_found:
            if format_found not in pools:
                pools[format_found] = []
            pools[format_found].append(col)
        else:
            # Unknown format, default to Standard
            if "Standard" not in pools:
                pools["Standard"] = []
            pools["Standard"].append(col)
    
    return pools


def split_row_by_units(row: pd.Series, units: float) -> List[pd.Series]:
    """
    Split a row with units > 1 into multiple placement entries.
    
    If units > 1.0, creates:
    - full_chunks copies with units=1.0
    - one remainder copy if remainder > 0
    
    All copies maintain the same Title and WK# (they're just placement entries).
    
    Args:
        row: The row to split
        units: The total units for this row
    
    Returns:
        List of row copies (Series objects) to be placed
    """
    if units <= 1.0:
        return [row]
    
    full_chunks = int(math.floor(units))
    remainder = units - full_chunks
    
    result = []
    # Create full_chunks copies with units=1.0
    for _ in range(full_chunks):
        row_copy = row.copy()
        # Set __units to 1.0 for packing algorithm
        row_copy["__units"] = 1.0
        result.append(row_copy)
    
    # Add remainder if > 0
    if remainder > 1e-6:
        row_copy = row.copy()
        row_copy["__units"] = remainder
        result.append(row_copy)
    
    return result


def build_layout_by_pool(
    rows_for_week: pd.DataFrame,
    screen_pools: Dict[str, List[int]],
) -> Dict[int, List[pd.Series]]:
    """
    Build layout by format pool, packing pieces into their respective screen columns.
    
    Args:
        rows_for_week: DataFrame of booking rows for the week
        screen_pools: Dict mapping format_key -> list of column indices
    
    Returns:
        Dict mapping column_index -> list of piece rows placed in that column
    """
    if rows_for_week is None or rows_for_week.empty:
        return {}
    
    # Step 1: Explode rows into format-specific pieces
    all_pieces = []
    for _, row in rows_for_week.iterrows():
        pieces = explode_row_into_pieces(row)
        all_pieces.extend(pieces)
    
    if not all_pieces:
        return {}
    
    # Step 2: Split pieces with units > 1 into full chunks + remainder
    expanded_pieces = []
    for piece in all_pieces:
        units = float(piece["__units"])
        full = int(math.floor(units))
        rem = units - full
        
        # Emit full chunks with units = 1.0
        for _ in range(full):
            piece_copy = piece.copy()
            piece_copy["__units"] = 1.0
            expanded_pieces.append(piece_copy)
        
        # Emit remainder if > 0
        if rem > 0.0001:
            piece_copy = piece.copy()
            piece_copy["__units"] = rem
            expanded_pieces.append(piece_copy)
    
    # Step 3: Calculate metadata for sorting
    SPECIAL_UNIT_COLS = [col for col in SCREEN_UNIT_COLS if col.lower() != "standard"]
    
    for piece in expanded_pieces:
        # __is_special: True if format != "Standard"
        piece["__is_special"] = (piece.get("__format", "Standard") != "Standard")
        
        # __new_priority: 0 for NEW movies, 1 for others
        status = normalize_status(piece.get(COL_STATUS))
        piece["__new_priority"] = 0 if status.startswith("new") else 1
    
    # Step 4: Group pieces by format
    pieces_by_format: Dict[str, List[pd.Series]] = {}
    for piece in expanded_pieces:
        format_key = piece.get("__format", "Standard")
        if format_key not in pieces_by_format:
            pieces_by_format[format_key] = []
        pieces_by_format[format_key].append(piece)
    
    # Step 5: Pack each format group into its pool
    result: Dict[int, List[pd.Series]] = {}  # col_idx -> list of pieces
    
    for format_key, pieces in pieces_by_format.items():
        # Get screen columns for this format
        format_cols = screen_pools.get(format_key, [])
        
        # If no columns for this format, fallback to Standard
        if not format_cols:
            format_cols = screen_pools.get("Standard", [])
        
        # If still no columns, skip (or could use Unplayed later)
        if not format_cols:
            continue
        
        # Sort pieces for this format
        pieces_df = pd.DataFrame(pieces)
        pieces_df = pieces_df.sort_values(
            by=["__is_special", "__new_priority", "__units"],
            ascending=[True, True, False],
            kind="mergesort",
        )
        
        # First-fit packing into format columns
        epsilon = 1e-6
        screens: List[Dict[str, object]] = []  # Each screen: {"remaining": float, "pieces": List[pd.Series], "col_idx": int}
        
        # Initialize screens for this format
        for col_idx in format_cols:
            screens.append({
                "remaining": 1.0,
                "pieces": [],
                "col_idx": col_idx
            })
        
        # Pack pieces into screens
        for _, piece_row in pieces_df.iterrows():
            units = float(piece_row["__units"])
            placed = False
            
            # Try to place in existing screen with enough capacity
            for screen in screens:
                if screen["remaining"] >= units - epsilon:
                    screen["pieces"].append(piece_row)
                    screen["remaining"] = screen["remaining"] - units
                    placed = True
                    break
            
            # If not placed, try to create overflow (use last screen or extend)
            if not placed:
                # Use the last screen in the pool (will exceed capacity, but that's okay)
                if screens:
                    screens[-1]["pieces"].append(piece_row)
                    screens[-1]["remaining"] = max(0.0, screens[-1]["remaining"] - units)
        
        # Add packed pieces to result
        for screen in screens:
            col_idx = screen["col_idx"]
            if screen["pieces"]:
                if col_idx not in result:
                    result[col_idx] = []
                result[col_idx].extend(screen["pieces"])
    
    return result


def build_screen_layout(rows_for_week: pd.DataFrame) -> List[List[pd.Series]]:
    """
    Build a list of screens for the given week (legacy function for backward compatibility).

    Each screen is represented as a list of pandas Series objects in the
    order they should appear within the cell.
    
    NOTE: This function is kept for backward compatibility but should be replaced
    with build_layout_by_pool() for format-aware packing.
    """
    if rows_for_week is None or rows_for_week.empty:
        return []

    df = rows_for_week.copy()
    df["__units"] = df.apply(compute_units, axis=1)
    
    # Filter to rows with units > 0 FIRST (before splitting)
    df = df[df["__units"] > 0].copy()
    
    if df.empty:
        return []
    
    # Step 1: Expand rows into multiple "pieces" BEFORE sorting
    # Split rows with units > 1 into full chunks (1.0 each) + remainder (< 1.0)
    expanded_rows = []
    for _, row in df.iterrows():
        units = row["__units"]
        full = int(math.floor(units))
        rem = units - full
        
        # Emit `full` copies with piece_units = 1.0
        for _ in range(full):
            row_copy = row.copy()
            row_copy["__units"] = 1.0
            expanded_rows.append(row_copy)
        
        # If rem > 0.0001, emit one copy with piece_units = rem
        if rem > 0.0001:
            row_copy = row.copy()
            row_copy["__units"] = rem
            expanded_rows.append(row_copy)
    
    # Step 2: Calculate metadata for sorting (using original row data)
    # We need to preserve sorting metadata for each piece
    SPECIAL_UNIT_COLS = [col for col in SCREEN_UNIT_COLS if col.lower() != "standard"]
    
    for row_piece in expanded_rows:
        # Calculate __is_special: True if uses any premium format
        is_special = False
        for col in SPECIAL_UNIT_COLS:
            if col in row_piece.index:
                value = row_piece.get(col)
                if pd.notna(value) and float(value) > 0:
                    is_special = True
                    break
        row_piece["__is_special"] = is_special
        
        # Calculate __new_priority: 0 for NEW movies, 1 for others
        status = normalize_status(row_piece.get(COL_STATUS))
        row_piece["__new_priority"] = 0 if status.startswith("new") else 1
    
    # Step 3: Sort by special, new_priority, units (descending)
    # Convert to DataFrame for sorting (preserves Series structure when iterating)
    expanded_df = pd.DataFrame(expanded_rows)
    expanded_df = expanded_df.sort_values(
        by=["__is_special", "__new_priority", "__units"],
        ascending=[True, True, False],
        kind="mergesort",
    )
    
    # Step 4: Run "first fit" packing algorithm
    # Each piece now has __units <= 1.0
    epsilon = 1e-6
    screens: List[Dict[str, object]] = []  # Each screen: {"remaining": float, "movies": List[pd.Series]}

    for _, row_piece in expanded_df.iterrows():
        units = float(row_piece["__units"])
        placed = False
        
        # Try to place in existing screen with enough remaining capacity
        for screen in screens:
            if screen["remaining"] >= units - epsilon:
                screen["movies"].append(row_piece)
                screen["remaining"] = screen["remaining"] - units
                placed = True
                break
        
        # If not placed, create new screen
        if not placed:
            screens.append({
                "remaining": 1.0 - units,  # units is <= 1.0, so remaining >= 0
                "movies": [row_piece]
            })

    return [screen["movies"] for screen in screens]


def build_screen_layout_standard_then_atmos(
    rows_for_week: pd.DataFrame,
    extra_non_new_standard_screens: Optional[List[List[pd.Series]]] = None,
    return_block_info: bool = False,
) -> object:
    """
    Circuit-agnostic, sequential layout builder.

    Ordering (no special header labels):
      1) NEW pieces from the first detected unit column (STANDARD_UNIT_COL)
      2) Non-NEW pieces from STANDARD_UNIT_COL
      3) For each remaining detected unit column (PREMIUM_UNIT_COLS), pack that format
         into its own screens and append in file order (no sharing between formats)

    This preserves the existing NEW-vs-non-NEW separation behavior for the standard pool,
    while making premium pools general (ATMOS/MPX/DBOX/3D/etc.) without hardcoding names.
    """
    if rows_for_week is None or rows_for_week.empty:
        return []

    if not STANDARD_UNIT_COL:
        # If we couldn't infer any unit columns, fall back to empty layout.
        return []

    df = rows_for_week.copy()

    # --- unit calculator for a specific column ---
    def units_in_col(row: pd.Series, col_name: str) -> float:
        if col_name in row.index:
            v = row.get(col_name)
            if pd.notna(v):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return 0.0
        return 0.0

    # --- helpers ---
    def expand_rows(df_in: pd.DataFrame, col_name: str) -> List[pd.Series]:
        expanded: List[pd.Series] = []
        for _, row in df_in.iterrows():
            total_units = units_in_col(row, col_name)
            if total_units <= 0:
                continue

            # Split into whole screens + remainder so each piece has __units <= 1.0
            full = int(total_units)
            rem = total_units - full

            for _ in range(full):
                r = row.copy()
                r["__units"] = 1.0
                expanded.append(r)

            if rem > 1e-6:
                r = row.copy()
                r["__units"] = rem
                expanded.append(r)

        return expanded

    def pack_pieces(pieces: List[pd.Series]) -> List[List[pd.Series]]:
        if not pieces:
            return []

        # Larger pieces first so halves/quarters fill neatly
        pieces.sort(key=lambda x: float(x.get("__units", 0.0)), reverse=True)

        screens: List[Dict[str, object]] = []
        epsilon = 1e-6

        for row_piece in pieces:
            units = float(row_piece.get("__units", 0.0))
            placed = False

            for screen in screens:
                if screen["remaining"] >= units - epsilon:
                    screen["movies"].append(row_piece)
                    screen["remaining"] -= units
                    placed = True
                    break

            if not placed:
                screens.append({"remaining": 1.0 - units, "movies": [row_piece]})

        return [s["movies"] for s in screens]

    # ---- Pools ----
    def is_new_row(r: pd.Series) -> bool:
        return normalize_status(r.get(COL_STATUS)).startswith("new")

    # ---- Build layouts ----
    # Default behavior (Warehouse-safe): keep NEW and non-NEW packed separately.
    df_new = df[df.apply(is_new_row, axis=1)].copy()
    df_non_new = df[~df.apply(is_new_row, axis=1)].copy()

    new_std_layout_sep = pack_pieces(expand_rows(df_new, STANDARD_UNIT_COL))
    non_new_std_layout_sep = pack_pieces(expand_rows(df_non_new, STANDARD_UNIT_COL))
    sep_layout = new_std_layout_sep + non_new_std_layout_sep + (extra_non_new_standard_screens or [])
    sep_count = len(sep_layout)

    # Maya-style fix: allow fractional pieces to share across NEW/non-NEW ONLY if it reduces
    # the total screen count (prevents phantom extra screens like a 16-screen house becoming 17).
    df_std_sorted = df.copy()
    try:
        df_std_sorted["__is_new"] = df_std_sorted.apply(is_new_row, axis=1)
        df_std_sorted["__std_units"] = df_std_sorted.apply(lambda r: units_in_col(r, STANDARD_UNIT_COL), axis=1)
        df_std_sorted = df_std_sorted.sort_values(["__is_new", "__std_units"], ascending=[False, False])
    except Exception:
        pass

    std_layout = pack_pieces(expand_rows(df_std_sorted, STANDARD_UNIT_COL))

    new_std_layout_mix = [s for s in std_layout if any(is_new_row(rr) for rr in s)]
    non_new_std_layout_mix = [s for s in std_layout if not any(is_new_row(rr) for rr in s)]
    mix_layout = new_std_layout_mix + non_new_std_layout_mix + (extra_non_new_standard_screens or [])
    mix_count = len(mix_layout)

    if mix_count < sep_count:
        new_std_layout = new_std_layout_mix
        non_new_std_layout = non_new_std_layout_mix
    else:
        new_std_layout = new_std_layout_sep
        non_new_std_layout = non_new_std_layout_sep

    premium_layouts: List[List[pd.Series]] = []
    for prem_col in PREMIUM_UNIT_COLS:
        premium_layouts += pack_pieces(expand_rows(df, prem_col))

    combined_layout = new_std_layout + non_new_std_layout + (extra_non_new_standard_screens or []) + premium_layouts

    if return_block_info:
        # standard pool screen count (NEW + non-NEW + extras)
        standard_count = len(new_std_layout) + len(non_new_std_layout) + len(extra_non_new_standard_screens or [])
        # premium blocks screen counts, in the same order as PREMIUM_UNIT_COLS
        premium_blocks: List[Tuple[str, int]] = []
        # We already appended premium layouts to premium_layouts, but we also need per-format counts.
        # Re-pack per format to get stable counts without changing the combined output.
        for prem_col in PREMIUM_UNIT_COLS:
            prem_screens = pack_pieces(expand_rows(df, prem_col))
            premium_blocks.append((prem_col, len(prem_screens)))

        return combined_layout, {"standard_screens": standard_count, "premium_blocks": premium_blocks}

    return combined_layout


def find_host_screen_for_split(split_row: pd.Series, layout: List[List[pd.Series]]) -> Optional[int]:
    """
    Find the best host screen for a split/swipe row.
    
    Strategy:
    1. First choice: same distributor (DIST) match
       - If multiple DIST matches, pick the one with highest FSS
    2. If no DIST match: pick the screen with highest-FSS movie overall
    
    Args:
        split_row: The split/swipe row to find a host for
        layout: The current layout (list of screen movie lists)
    
    Returns:
        Index of the best host screen, or None if no screens exist
    """
    if not layout:
        return None
    
    split_dist = split_row.get(COL_DIST)
    split_dist_str = str(split_dist).strip() if pd.notna(split_dist) else None
    
    best_screen_idx = None
    best_fss = -1.0
    
    # Try to find DIST match first
    for screen_idx, screen_movies in enumerate(layout):
        for host_row in screen_movies:
            host_dist = host_row.get(COL_DIST)
            host_dist_str = str(host_dist).strip() if pd.notna(host_dist) else None
            
            # Check DIST match
            if split_dist_str and host_dist_str and split_dist_str.lower() == host_dist_str.lower():
                # Same distributor - check FSS
                host_fss = host_row.get(COL_FSS)
                if pd.notna(host_fss):
                    try:
                        fss_val = float(host_fss)
                        if fss_val > best_fss:
                            best_fss = fss_val
                            best_screen_idx = screen_idx
                    except (TypeError, ValueError):
                        pass
    
    # If we found a DIST match, use it
    if best_screen_idx is not None:
        return best_screen_idx
    
    # No DIST match - find highest FSS overall
    best_fss = -1.0
    for screen_idx, screen_movies in enumerate(layout):
        for host_row in screen_movies:
            host_fss = host_row.get(COL_FSS)
            if pd.notna(host_fss):
                try:
                    fss_val = float(host_fss)
                    if fss_val > best_fss:
                        best_fss = fss_val
                        best_screen_idx = screen_idx
                except (TypeError, ValueError):
                    pass
    
    # If still no match, use first screen as fallback
    return best_screen_idx if best_screen_idx is not None else 0


def to_excel_datetime(value: pd.Timestamp | datetime | date | None) -> Optional[datetime]:
    """Convert pandas timestamps or date objects to datetime for Excel."""
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raise TypeError(f"Unsupported date type: {type(value)}")


def find_date_row(ws: Worksheet, target_date: datetime, start_row: int = 3) -> Optional[int]:
    """
    Return the row index where column A matches the target date.
    
    Args:
        ws: Worksheet to search
        target_date: Date to find
        start_row: Row to start searching from (default 3, after headers)
    """
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, datetime):
            if value.date() == target_date.date():
                return row
        elif isinstance(value, date):
            if value == target_date.date():
                return row
        elif isinstance(value, str):
            try:
                parsed = pd.to_datetime(value, errors="coerce")
                if pd.notna(parsed) and parsed.date() == target_date.date():
                    return row
            except Exception:
                continue
    return None


def ensure_week_rows(ws: Worksheet, play_week: datetime, start_row: int = 3) -> Tuple[int, int]:
    """
    Ensure the worksheet has rows for the play week date and its FSS row.
    
    Weeks are written starting at start_row (default 3, after headers in rows 1-2).

    Returns a tuple of (titles_row_index, fss_row_index).
    """
    date_row = find_date_row(ws, play_week, start_row=start_row)
    if date_row is None:
        # Append after the last existing row, but ensure we start at least at start_row
        if ws.max_row < start_row:
            date_row = start_row
        else:
            date_row = ws.max_row + 1
        date_cell = ws.cell(row=date_row, column=1, value=play_week)
        date_cell.number_format = "m/d/yy"
        date_cell.alignment = WRAP_TOP
        # Ensure there's a row for FSS directly below.
        fss_row = date_row + 1
        ws.cell(row=fss_row, column=1, value=None)
    else:
        fss_row = date_row + 1

    # Ensure consistent date formatting/alignment for the week row
    date_cell = ws.cell(row=date_row, column=1)
    date_cell.number_format = "m/d/yy"
    date_cell.alignment = WRAP_TOP

    return date_row, fss_row


def format_title_line(row: pd.Series) -> str:
    """Create the display string for a movie title line."""
    title = row.get(COL_TITLE, "Untitled")
    week_num = row.get(COL_WKNUM)
    if pd.notna(week_num):
        try:
            week_int = int(week_num)
            result = f"{title} - wk{week_int}"
        except (TypeError, ValueError):
            result = f"{title} - wk{week_num}"
    else:
        result = str(title)
    # Sanitize before returning
    return sanitize_excel_text(result)



def _normalize_for_match(s: str) -> str:
    return re.sub(r"[^a-z0-9&]+", " ", str(s).lower()).strip()

_STOPWORDS = {"the", "a", "an", "of", "and"}

def _title_aliases(title: str) -> set[str]:
    """
    Alias set for matching titles referenced in Comments.
    Includes:
      - normalized full title
      - first word, first two words
      - initials (e.g., "tt")
      - "t&t" style for 2-word initialisms
      - version without digits (helps "Black Phone" match "Black Phone 2")
    """
    t = str(title or "").strip()
    if not t:
        return set()

    norm = _normalize_for_match(t)
    tokens = [tok for tok in norm.split() if tok and tok not in _STOPWORDS]
    aliases: set[str] = set()

    aliases.add(norm)
    if len(tokens) >= 1:
        aliases.add(tokens[0])
    if len(tokens) >= 2:
        aliases.add(" ".join(tokens[:2]))

    initials = "".join(tok[0] for tok in tokens if tok and tok[0].isalnum())
    if initials:
        aliases.add(initials)
        if len(initials) == 2:
            aliases.add(f"{initials[0]}&{initials[1]}")  # "t&t"

    nodigits = re.sub(r"\d+", "", norm).strip()
    if nodigits:
        aliases.add(nodigits)

    return {a for a in aliases if a}

def _mutual_star_holds(star_df: pd.DataFrame) -> pd.DataFrame:
    """
    From 'Hold * shows' rows, select the best mutually-linked pair(s) based on Comments.
    We intentionally keep this conservative so these rows annotate an existing STANDARD
    screen (e.g., the Black Phone triple-feature) without pulling in unrelated one-offs.

    Selection rule:
      1) Build mutual mention pairs A<->B using title alias matching in Comments.
      2) Choose the pair with the lowest max(WK#) (avoids long-run placeholders like wk7).
         Tie-breaker: highest combined FSS.
      3) Return only the rows in the chosen pair. If no mutual pairs, return empty.
    """
    if star_df is None or star_df.empty:
        return star_df

    # Normalize titles + build aliases
    titles = star_df[COL_TITLE].fillna("").astype(str).tolist()
    title_norms = [_normalize_for_match(t) for t in titles]
    alias_map = {tn: _title_aliases(titles[i]) for i, tn in enumerate(title_norms)}

    # Comments
    comments = star_df.get("Comments", pd.Series([""] * len(star_df))).fillna("").astype(str).tolist()

    # Directed mentions: A -> {B...}
    mention: dict[str, set[str]] = {tn: set() for tn in title_norms}
    for i, a_norm in enumerate(title_norms):
        c = _normalize_for_match(comments[i])
        if not c:
            continue
        for b_norm in title_norms:
            if b_norm == a_norm:
                continue
            if any(alias in c for alias in alias_map[b_norm]):
                mention[a_norm].add(b_norm)

    # Helper to read WK# / FSS
    def _wk_of(norm_title: str) -> float:
        # pick first matching row
        idx = title_norms.index(norm_title)
        v = star_df.iloc[idx].get(COL_WKNUM, None)
        try:
            return float(v)
        except Exception:
            return float("inf")

    def _fss_of(norm_title: str) -> float:
        idx = title_norms.index(norm_title)
        v = star_df.iloc[idx].get(COL_FSS, None)
        try:
            return float(v)
        except Exception:
            return 0.0

    # Collect mutual pairs (a < b to dedupe)
    pairs = []
    for a in title_norms:
        for b in mention[a]:
            if a in mention.get(b, set()):
                if a < b:
                    pairs.append((a, b))

    if not pairs:
        return star_df.iloc[0:0].copy()

    # Score pairs
    best_pair = None
    best_key = None
    for a, b in pairs:
        key = (
            max(_wk_of(a), _wk_of(b)),      # prefer smaller max WK#
            -(_fss_of(a) + _fss_of(b)),     # then higher combined FSS
            a, b
        )
        if best_key is None or key < best_key:
            best_key = key
            best_pair = (a, b)

    keep_norm = set(best_pair)
    mask = [tn in keep_norm for tn in title_norms]
    return star_df.loc[mask].copy()

def _attach_rows_to_standard_screens(
    layout: List[List[pd.Series]],
    attach_df: pd.DataFrame,
) -> List[List[pd.Series]]:
    """
    Attach rows into existing STANDARD screens based on comment matching.
    Does NOT create new screens and does NOT change screen counts.
    """
    if attach_df is None or attach_df.empty or not layout:
        return layout

    # Mark premium screens (any row with ATMOS units > 0)
    premium_flags: list[bool] = []
    for screen in layout:
        is_premium = False
        for r in screen:
            try:
                v = r.get(COL_ATMOS, 0)
                if pd.notna(v) and float(v) > 1e-6:
                    is_premium = True
                    break
            except Exception:
                pass
        premium_flags.append(is_premium)

    # Alias bag for each STANDARD screen
    screen_aliases: list[set[str]] = []
    for idx, screen in enumerate(layout):
        if premium_flags[idx]:
            screen_aliases.append(set())
            continue
        aliases: set[str] = set()
        for r in screen:
            aliases |= _title_aliases(r.get(COL_TITLE, ""))
        screen_aliases.append(aliases)

    for _, row in attach_df.iterrows():
        comment = _normalize_for_match(row.get("Comments", ""))
        if not comment:
            continue

        best_idx = None
        best_score = 0
        for idx, aliases in enumerate(screen_aliases):
            if premium_flags[idx]:
                continue
            score = 0
            for a in aliases:
                if a and a in comment:
                    score += 3 if " " in a else 1
            if score > best_score:
                best_score = score
                best_idx = idx

        # Require a meaningful match to avoid random attachments
        if best_idx is not None and best_score >= 2:
            layout[best_idx].append(row)

    return layout



def _build_star_hold_addon_screens(
    star_hold_rows: pd.DataFrame,
    host_rows: pd.DataFrame,
) -> List[List[pd.Series]]:
    """
    Turn 'Hold * shows' fractional STANDARD rows into *additional* STANDARD screens.

    Why this exists:
      - These rows have real fractional Standard units (0.25/0.5/etc.) and must be
        represented as their own standard screen block(s), which in turn pushes
        ATMOS/premium screens to the right (e.g., ATMOS should land in screen 10).

    Output format per screen:
      [host_row_copy, addon_row_1, addon_row_2, ...]
    Host selection:
      - Uses comment matching against title aliases from host_rows.
      - If no host match, the addon row is scheduled as a standalone screen (no host line).

    IMPORTANT: This does not change how NEW/ATMOS/standard packing works for the main rows.
    It only adds missing standard multi-feature blocks as extra screens.
    """
    if star_hold_rows is None or star_hold_rows.empty:
        return []

    # Build alias bags for host candidates (STANDARD-only host rows)
    host_candidates: list[pd.Series] = []
    host_aliases: list[set[str]] = []

    if host_rows is not None and not host_rows.empty:
        tmp = host_rows.copy()
        # Only hosts that actually have STANDARD demand (so "extra ___" attaches to real standard titles)
        std = pd.to_numeric(tmp.get(COL_STANDARD), errors="coerce").fillna(0.0)
        tmp = tmp[std > 1e-6].copy()

        for _, r in tmp.iterrows():
            host_candidates.append(r)
            host_aliases.append(_title_aliases(r.get(COL_TITLE, "")))

    # Assign each star-hold row to a host index (or None)
    groups: dict[Optional[int], list[pd.Series]] = {}

    for _, r in star_hold_rows.iterrows():
        comment = _normalize_for_match(r.get("Comments", ""))
        best_idx: Optional[int] = None
        best_score = 0

        if comment and host_candidates:
            for idx, aliases in enumerate(host_aliases):
                score = 0
                for a in aliases:
                    if a and a in comment:
                        score += 3 if " " in a else 1
                if score > best_score:
                    best_score = score
                    best_idx = idx

            # Require a meaningful match; otherwise treat as standalone
            if best_score < 2:
                best_idx = None

        groups.setdefault(best_idx, []).append(r)

    # Bin-pack the addon rows per host into 1.0 screens using their STANDARD units
    def addon_units(row: pd.Series) -> float:
        v = row.get(COL_STANDARD)
        try:
            return float(v) if pd.notna(v) else 0.0
        except Exception:
            return 0.0

    def pack_addons(addons: list[pd.Series]) -> list[list[pd.Series]]:
        pieces = []
        for a in addons:
            u = addon_units(a)
            if u > 1e-6:
                aa = a.copy()
                aa["__units"] = u
                pieces.append(aa)
        if not pieces:
            return []

        # Larger pieces first so 0.5 then 0.25s fill neatly
        pieces.sort(key=lambda x: float(x.get("__units", 0.0)), reverse=True)

        screens: list[dict[str, object]] = []
        epsilon = 1e-6

        for p in pieces:
            u = float(p.get("__units", 0.0))
            placed = False
            for s in screens:
                if s["remaining"] >= u - epsilon:
                    s["movies"].append(p)
                    s["remaining"] -= u
                    placed = True
                    break
            if not placed:
                screens.append({"remaining": 1.0 - u, "movies": [p]})

        return [s["movies"] for s in screens]

    out_screens: list[list[pd.Series]] = []

    # Deterministic ordering: hosts first by title, then standalone group last
    for host_idx in sorted([k for k in groups.keys() if k is not None]):
        addons = groups.get(host_idx, [])
        host_row = host_candidates[host_idx].copy()

        # Only prepend the host title if the host actually consumes >1 STANDARD screen.
        # If the host is a single-screen title (Standard ~= 1.0), we do NOT duplicate it
        # in the addon multi-feature screen (this was causing Black Phone 2 to appear twice).
        host_std = pd.to_numeric(host_row.get(COL_STANDARD), errors="coerce")
        host_std = float(host_std) if pd.notna(host_std) else 0.0
        eps = 1e-6

        for addon_screen in pack_addons(addons):
            if host_std > 1.0 + eps:
                # Put host line first, then the addon lines
                out_screens.append([host_row.copy()] + addon_screen)
            else:
                # Single-screen host: addon screen contains only the addon lines
                out_screens.append(addon_screen)

    # Standalone group (no host match): each packed screen is just those addon rows
    if None in groups:
        for addon_screen in pack_addons(groups[None]):
            out_screens.append(addon_screen)

    return out_screens


def format_fss_value(value: object, is_final: bool = False) -> str:
    """
    Format the FSS number for display within the cell.
    
    Args:
        value: The FSS value (number or string)
        is_final: If True, append " (Final)" to the formatted value
    """
    if value is None:
        return ""
    if isinstance(value, str) and value.strip() == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if pd.isna(number) or (isinstance(number, float) and math.isnan(number)):
        return ""
    # Format with commas, no decimals (round if needed)
    if abs(number - round(number)) < 0.5:
        formatted = f"{int(round(number)):,}"
    else:
        formatted = f"{number:,.0f}"
    if is_final:
        formatted += " (Final)"
    # Sanitize before returning
    return sanitize_excel_text(formatted)


TITLE_WK_RE = re.compile(r"\s*-\s*wk.*$", re.IGNORECASE)
NON_ALNUM = re.compile(r"[^a-z0-9]+")

# Column alias mappings (case-insensitive matching)
COLUMN_ALIASES: Dict[str, List[str]] = {
    "Circuit": ["Circuit", "Chain", "Company", "Exhibitor"],
    "Theatre Name": ["Theatre Name", "Theater Name", "Location", "Site", "Venue"],
    "City": ["City"],
    "ST": ["ST", "State", "Province"],
    "Title": ["Title"],
    "Status": ["Status", "Booking Status", "Run Status"],
    "Playwk": ["Playwk", "Play Week", "Week Start", "Play Week Start", "WeekOf"],
    "WK#": ["WK#", "WK", "Week#", "Week #", "WeekNum", "Week Num"],
    "FSS": ["FSS", "FriSatSun", "WeekendGross", "Weekend"],
    "Total": ["Total"],
    "DIST": ["DIST", "Distributor", "Dist."],
}

# Required canonical columns
REQUIRED_COLUMNS = [
    "Circuit", "Theatre Name", "City", "ST", "Title", "Status", "Playwk", "WK#", "FSS"
]


def normalize_title_key(s: str) -> str:
    """
    Normalize a title string to a key for matching.
    
    Aggressively normalizes titles on both sides (bookings and sheet) for robust matching.
    
    - Remove trailing "- wk<number>" suffix
    - Convert to lowercase
    - Normalize punctuation/hyphens/commas/colons to spaces
    - Collapse whitespace
    """
    s = (s or "").strip()
    # Remove trailing "- wk<number>"
    s = re.sub(r"\s*-\s*wk\s*\d+\s*$", "", s, flags=re.IGNORECASE)
    s = s.lower()
    # Normalize punctuation/hyphens/commas/colons
    s = NON_ALNUM.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def title_key_from_line(title_line: str) -> str:
    """Normalize a title line like 'Movie Name - wk3' -> 'movie name'."""
    return normalize_title_key(title_line)


def split_cell_movies(cell_text: object) -> list[str]:
    """
    Split a titles cell into separate movie title lines.
    Titles are separated by blank lines (double newline) or single newlines.
    Returns a list of title lines (first line of each block).
    """
    if cell_text is None:
        return []
    text = str(cell_text).replace("\r\n", "\n").strip()
    if not text:
        return []
    # Prefer double-newline blocks, but fallback to single-newline lines
    if "\n\n" in text:
        blocks = re.split(r"\n\s*\n", text)
        out = []
        for b in blocks:
            line = b.strip().splitlines()[0].strip() if b.strip() else ""
            if line:
                out.append(line)
        return out
    # Fallback: treat each non-empty line as a title line
    return [ln.strip() for ln in text.split("\n") if ln.strip()]


def write_prev_week_fss_from_existing_titles(
    ws: Worksheet,
    titles_row: int,
    fss_row: int,
    screen_cols: list[int],
    fss_map: dict[str, str],
) -> None:
    """
    Write FSS to prev week row by reading the existing titles row.
    This preserves stacked layouts exactly.
    Uses rich text so only Final lines are red (not the whole cell).
    """
    for col in screen_cols:
        titles_text = ws.cell(row=titles_row, column=col).value
        title_lines = split_cell_movies(titles_text)
        if not title_lines:
            fss_cell = ws.cell(row=fss_row, column=col, value=None)
            fss_cell.alignment = WRAP_TOP
            fss_cell.border = GRID_BORDER
            continue

        fss_lines = []

        for title_line in title_lines:
            key = normalize_title_key(title_line)
            fss_val = fss_map.get(key, "")
            
            # Fallback: partial match if exact key fails
            if not fss_val:
                for k, v in fss_map.items():
                    if key in k or k in key:
                        fss_val = v
                        break
            
            if fss_val:
                # Sanitize fss_val before appending
                fss_val = sanitize_excel_text(fss_val)
                fss_lines.append(fss_val)

        fss_cell = ws.cell(row=fss_row, column=col)
        set_rich_text_multiline(fss_cell, fss_lines)
        fss_cell.alignment = WRAP_TOP
        fss_cell.border = GRID_BORDER


def contains_status(rows: Sequence[pd.Series], keyword: str) -> bool:
    """Return True if any row status contains the keyword."""
    keyword = keyword.lower()
    for row in rows:
        status = normalize_status(row.get(COL_STATUS))
        if keyword in status:
            return True
    return False


def has_new_movie(rows: Sequence[pd.Series]) -> bool:
    """Return True if any row represents a new movie."""
    for row in rows:
        status = normalize_status(row.get(COL_STATUS))
        if status.startswith("new"):
            return True
    return False


def write_screen_cells(
    ws: Worksheet,
    titles_row: int,
    fss_row: Optional[int],
    layout: Union[List[List[pd.Series]], Dict[int, List[pd.Series]]],
    screen_cols: List[int],
    unplayed_col: Optional[int] = None,
    highlight_new: bool = False,
    screen_fill_by_idx: Optional[Dict[int, PatternFill]] = None,
) -> None:
    """
    Write layout data into worksheet screen columns.
    
    Supports both legacy format (List[List[pd.Series]]) and new pool-based format (Dict[int, List[pd.Series]]).
    
    Args:
        ws: Worksheet to write to
        titles_row: Row number for titles
        fss_row: Row number for FSS (or None)
        layout: Either:
            - List of screen layouts (each is a list of movie rows) - legacy format
            - Dict mapping column_index -> list of piece rows - new pool-based format
        screen_cols: List of screen column indices (1-based)
        unplayed_col: Column index for overflow movies (or None)
        highlight_new: Whether to highlight new movies
        screen_fill_by_idx: Optional dict mapping screen index to fill color for new movies
    """
    # Check if layout is dict-based (new format) or list-based (legacy)
    if isinstance(layout, dict):
        # New pool-based format: layout is Dict[col_idx, list of pieces]
        for col_idx, pieces in layout.items():
            if pieces:
                # Stack titles with double newline between them
                title_lines = [format_title_line(piece) for piece in pieces]
                titles = "\n\n".join(title_lines)
                # Sanitize the final titles string before writing
                titles = sanitize_excel_text(titles)
                titles_cell = ws.cell(row=titles_row, column=col_idx, value=titles or None)
                titles_cell.alignment = WRAP_TOP
                titles_cell.border = GRID_BORDER

                # Apply new movie highlighting if specified
                if highlight_new:
                    if screen_fill_by_idx and col_idx in screen_fill_by_idx:
                        # Use provided fill (yellow for first occurrence, orange for repeats)
                        titles_cell.fill = screen_fill_by_idx[col_idx]
                    elif has_new_movie(pieces):
                        # Fallback: use yellow if no specific fill provided
                        titles_cell.fill = NEW_YELLOW
                    if has_new_movie(pieces):
                        titles_cell.font = Font(bold=True)

                if fss_row is not None:
                    # Format FSS with (Final) suffix where appropriate
                    fss_lines = []
                    for piece in pieces:
                        is_final = "final" in normalize_status(piece.get(COL_STATUS))
                        fss_val = format_fss_value(piece.get(COL_FSS), is_final=is_final)
                        if fss_val:
                            fss_lines.append(fss_val)
                    fss_cell = ws.cell(row=fss_row, column=col_idx)
                    set_rich_text_multiline(fss_cell, fss_lines)
                    fss_cell.alignment = WRAP_TOP
                    fss_cell.border = GRID_BORDER
    else:
        # Legacy list-based format
        for idx, screen_rows in enumerate(layout):
            if idx < len(screen_cols):
                column = screen_cols[idx]
                # Stack titles with double newline between them
                title_lines = [format_title_line(row) for row in screen_rows]
                titles = "\n\n".join(title_lines)
                # Sanitize the final titles string before writing
                titles = sanitize_excel_text(titles)
                titles_cell = ws.cell(row=titles_row, column=column, value=titles or None)
                titles_cell.alignment = WRAP_TOP
                titles_cell.border = GRID_BORDER

                # Apply new movie highlighting if specified (only for current week)
                if highlight_new:
                    if screen_fill_by_idx and idx in screen_fill_by_idx:
                        titles_cell.fill = screen_fill_by_idx[idx]
                    elif has_new_movie(screen_rows):
                        # Fallback: use yellow if no specific fill provided
                        titles_cell.fill = NEW_YELLOW
                    if has_new_movie(screen_rows):
                        titles_cell.font = Font(bold=True)
                # Note: Do NOT remove fills on non-new writes - preserve historical formatting

                if fss_row is not None:
                    # Format FSS with (Final) suffix where appropriate
                    fss_lines = []
                    for row in screen_rows:
                        is_final = "final" in normalize_status(row.get(COL_STATUS))
                        fss_val = format_fss_value(row.get(COL_FSS), is_final=is_final)
                        if fss_val:
                            fss_lines.append(fss_val)
                    fss_cell = ws.cell(row=fss_row, column=column)
                    set_rich_text_multiline(fss_cell, fss_lines)
                    fss_cell.alignment = WRAP_TOP
                    fss_cell.border = GRID_BORDER
        
        # Handle overflow: movies beyond available screens go to Unplayed column
        if unplayed_col is not None and len(layout) > len(screen_cols):
            overflow_movies = []
            for screen_rows in layout[len(screen_cols):]:
                for row in screen_rows:
                    overflow_movies.append(format_title_line(row))
            
            if overflow_movies:
                # Stack overflow titles with double newline
                overflow_text = "\n\n".join(overflow_movies)
                overflow_text = sanitize_excel_text(overflow_text)
                overflow_cell = ws.cell(row=titles_row, column=unplayed_col, value=overflow_text or None)
                overflow_cell.alignment = WRAP_TOP
                overflow_cell.border = GRID_BORDER


def clear_row_range(
    ws: Worksheet,
    row_idx: int,
    start_col: int = 2,
    end_col: Optional[int] = None,
    clear_fill: bool = True,
    clear_font: bool = True,
    clear_border: bool = False,
) -> None:
    """
    Clear cell values from start_col through end_col (or max_column) on a row.
    
    Args:
        ws: Worksheet to clear
        row_idx: Row number to clear
        start_col: Starting column (1-based)
        end_col: Ending column (1-based, or None for max_column)
        clear_fill: Whether to clear cell fill (default True)
        clear_font: Whether to reset font (default True)
        clear_border: Whether to clear borders (default False, borders are persistent)
    """
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.value = None
        if clear_fill:
            cell.fill = PatternFill()
        if clear_font:
            cell.font = Font()
        # Note: borders are never cleared (persistent)


def update_theater_sheet(
    ws: Worksheet,
    prev_layout: List[List[pd.Series]],
    curr_layout: Union[List[List[pd.Series]], Dict[int, List[pd.Series]]],
    prev_week: Optional[pd.Timestamp],
    curr_week: pd.Timestamp,
    screen_cols: List[int],
    unplayed_col: Optional[int] = None,
    prev_fss_map: Optional[Dict[str, str]] = None,
    alt_content_col: Optional[int] = None,
    screen_fill_by_idx: Optional[Dict[int, PatternFill]] = None,
) -> None:
    """
    Update a theater worksheet with previous and current week data.

    prev_layout is used to populate titles and FSS for the previous week.
    curr_layout populates the titles row for the current week.
    """
    # Build list of columns to style (screen columns + special columns)
    cols_to_style = list(screen_cols)
    if unplayed_col is not None:
        cols_to_style.append(unplayed_col)
    if alt_content_col is not None:
        cols_to_style.append(alt_content_col)
    
    if prev_week is not None:
        prev_dt = to_excel_datetime(prev_week)
        titles_row, fss_row = ensure_week_rows(ws, prev_dt)

        # IMPORTANT: do NOT clear or rewrite the titles row for prev week
        # Only clear FSS row values (but preserve fill for week banding)
        clear_row_range(ws, fss_row, clear_fill=False, clear_font=True)

        if prev_fss_map is None:
            prev_fss_map = {}

        write_prev_week_fss_from_existing_titles(
            ws=ws,
            titles_row=titles_row,
            fss_row=fss_row,
            screen_cols=screen_cols,
            fss_map=prev_fss_map,
        )
        
        # Apply week banding for previous week
        if prev_dt is not None:
            apply_week_band(
                ws, titles_row, fss_row, screen_cols,
                alt_content_col=alt_content_col,
                unplayed_col=unplayed_col,
                play_week=prev_dt,
            )
        
        # Apply borders to previous week block (including titles row to ensure borders exist)
        apply_borders_to_week_block(ws, titles_row, fss_row, cols_to_style)
        apply_border_to_date_col(ws, titles_row, fss_row)

    curr_dt = to_excel_datetime(curr_week)
    curr_titles_row, curr_fss_row = ensure_week_rows(ws, curr_dt)
    # Clear both rows for the current week (clear everything for fresh write)
    clear_row_range(ws, curr_titles_row)
    clear_row_range(ws, curr_fss_row)
    
    # Apply week banding for current week (BEFORE writing cells, so new movie fills can override)
    if curr_dt is not None:
        apply_week_band(
            ws, curr_titles_row, curr_fss_row, screen_cols,
            alt_content_col=alt_content_col,
            unplayed_col=unplayed_col,
            play_week=curr_dt,
        )
    
    # For the current week, write titles (no FSS) with new movie highlighting
    write_screen_cells(
        ws, curr_titles_row, None, curr_layout, screen_cols, unplayed_col,
        highlight_new=True, screen_fill_by_idx=screen_fill_by_idx
    )
    
    # Apply borders to current week block
    apply_borders_to_week_block(ws, curr_titles_row, curr_fss_row, cols_to_style)
    apply_border_to_date_col(ws, curr_titles_row, curr_fss_row)


def build_layout_rows(
    df: pd.DataFrame,
    mask: pd.Series,
) -> pd.DataFrame:
    """Helper to filter rows using mask and ensure units > 0."""
    if df is None or mask is None:
        return pd.DataFrame()
    subset = df[mask].copy()
    subset["__units"] = subset.apply(compute_units, axis=1)
    subset = subset[subset["__units"] > 0]
    return subset


def get_theaters(bookings_df: pd.DataFrame) -> pd.DataFrame:
    """
    Return a DataFrame of unique theaters from the bookings data.

    Each row represents one theater (Circuit + Theatre Name + City + ST).
    """
    return (
        bookings_df[["Circuit", "Theatre Name", "City", "ST"]]
        .dropna(subset=["Circuit", "Theatre Name", "City", "ST"])
        .drop_duplicates()
    )


def compute_theater_id(circuit: str, theater_name: str, city: str, state: str) -> str:
    """
    Compute a stable theater_id from theater attributes.
    
    Format: "{circuit}|{theatre_name}|{city}|{state}" (exact raw values, stripped)
    This ID is used to identify sheets across runs, regardless of sheet title changes.
    
    Args:
        circuit: Circuit name
        theater_name: Theatre name
        city: City name
        state: State code
    
    Returns:
        Stable theater_id string
    """
    return f"{str(circuit).strip()}|{str(theater_name).strip()}|{str(city).strip()}|{str(state).strip()}"


def find_sheet_by_theater_id(wb: Workbook, theater_id: str) -> Optional[Worksheet]:
    """
    Find a worksheet by its theater_id stored in META_CELL.
    
    Args:
        wb: Workbook to search
        theater_id: Theater ID to find
    
    Returns:
        Worksheet with matching theater_id, or None if not found
    """
    for ws in wb.worksheets:
        try:
            meta_value = ws[META_CELL].value
            if meta_value and str(meta_value).strip() == theater_id:
                return ws
        except (AttributeError, KeyError):
            continue
    return None


def normalize_sheet_key(name: str) -> str:
    """
    Normalize a sheet name to a key for matching purposes.
    
    Removes suffixes like " (2)", trailing ", <City>, <ST>" or ", <ST>",
    collapses whitespace, and strips punctuation for comparison.
    
    Args:
        name: Sheet name to normalize
    
    Returns:
        Normalized key string for comparison
    """
    if not name:
        return ""
    
    # Lowercase
    key = name.lower()
    
    # Remove suffix like " (2)", " (3)" at end
    key = re.sub(r'\s*\(\d+\)\s*$', '', key)
    
    # Remove trailing ", <City>, <ST>" or ", <ST>" pattern
    # Pattern: ", <word(s)>, <2-letter-state>" or ", <2-letter-state>"
    key = re.sub(r',\s*[^,]+,\s*[a-z]{2}\s*$', '', key)  # ", City, ST"
    key = re.sub(r',\s*[a-z]{2}\s*$', '', key)  # ", ST"
    
    # Collapse whitespace and strip
    key = re.sub(r'\s+', ' ', key).strip()

    # Strip trailing punctuation like commas/spaces created by 31-char truncation
    key = key.rstrip(' ,')
    
    return key


def normalize_theatre_name(theater_name: str, prefix_regex: Optional[re.Pattern] = None) -> str:
    """
    Normalize a theater name by optionally removing prefixes and cleaning whitespace.
    
    Args:
        theater_name: Original theater name
        prefix_regex: Optional compiled regex pattern to strip prefixes
    
    Returns:
        Normalized theater name
    
    Examples:
        - With prefix_regex=r"^(warehouse|wc)\\s*": "Warehouse Rotunda" -> "Rotunda"
        - Without prefix_regex: "Warehouse Rotunda" -> "Warehouse Rotunda" (only whitespace cleaned)
    """
    s = str(theater_name or "").strip()
    
    # Apply prefix regex if provided
    if prefix_regex is not None:
        s = prefix_regex.sub("", s).strip()
    
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s)
    # Remove trailing commas/spaces
    s = s.rstrip(" ,")
    return s


def map_theater_to_sheet_name(
    theater_name: str,
    state: str,
    city: Optional[str] = None,
    prefix_regex: Optional[re.Pattern] = None,
    sheet_name_format: str = "{theatre}, {state}",
    include_city: bool = False,
) -> str:
    """
    Map theater name to workbook sheet name.
    
    Args:
        theater_name: Original theater name
        state: State code
        city: Optional city name (for disambiguation)
        prefix_regex: Optional regex to strip prefixes from theater name
        sheet_name_format: Format string for sheet name (default: "{theatre}, {state}")
        include_city: If True, format becomes "{theatre} ({city}), {state}"
    
    Returns:
        Sheet name string
    
    Examples:
        - Default: "Rotunda" + "MD" -> "Rotunda, MD"
        - With city: "Rotunda" + "MD" + city="Hagerstown" + include_city=True -> "Rotunda (Hagerstown), MD"
    """
    normalized = normalize_theatre_name(theater_name, prefix_regex)
    
    if include_city and city:
        return f"{normalized} ({city}), {state}"
    else:
        return sheet_name_format.format(theatre=normalized, state=state, city=city or "")


def sanitize_filename(name: str) -> str:
    """Replace characters that aren't allowed in file names."""
    return re.sub(r'[\\/*?:"<>|]', "_", name)


def make_theater_filename(output_dir: str, circuit: str, theater_name: str, city: str, state: str) -> str:
    """Build a safe filename for a theater's Excel workbook (legacy function)."""
    base = f"{circuit} - {theater_name} - {city} {state}".strip()
    safe = sanitize_filename(base)
    return os.path.join(output_dir, safe + ".xlsx")


def ensure_screen_headers(ws: Worksheet, required_screens: int, header_row: int = 2) -> None:
    """
    Ensure worksheet has enough screen header columns, expanding if needed.
    
    Detects existing screen headers (pure integers OR strings starting with integer like "10 Atmos").
    Uses max_screen_num + 1 for next screen number to avoid duplicates.
    If existing < required_screens, inserts new columns before notes boundary or appends at end.
    Never overwrites existing non-empty headers.
    Styles new header cells like existing headers (HEADER_BLUE, bold, centered, border).
    
    Args:
        ws: Worksheet to update
        required_screens: Minimum number of screen columns needed
        header_row: Row number containing headers (default 2)
    """
    if header_row > ws.max_row:
        # No header row exists, create it
        ws.cell(row=header_row, column=1, value=None)  # Column A is date column
    
    # Regex patterns for screen headers
    pure_int_pattern = re.compile(r'^\s*\d+\s*$')  # Pure integer: "1", "2", "10", etc.
    starts_with_int_pattern = re.compile(r'^\s*(\d+)\b')  # Starts with integer: "10 Atmos", "7 IMAX", etc.
    
    # Find notes boundary (first column containing "/")
    notes_boundary_col: Optional[int] = None
    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is not None:
            cell_str = str(cell_value).strip()
            if "/" in cell_str:
                notes_boundary_col = col
                break
    
    # Find rightmost non-empty header cell
    rightmost_col = 2
    for col in range(2, ws.max_column + 1):
        if notes_boundary_col is not None and col >= notes_boundary_col:
            break
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is not None and (not isinstance(cell_value, str) or cell_value.strip() != ""):
            rightmost_col = col
    
    # Scan to detect existing screens and find max_screen_num
    existing_screens = []
    max_screen_num = 0
    scan_limit = notes_boundary_col if notes_boundary_col is not None else (rightmost_col + 1)
    
    for col in range(2, scan_limit):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is None:
            continue
        
        cell_str = str(cell_value).strip()
        
        # Check if it's a pure integer
        if pure_int_pattern.match(cell_str):
            try:
                screen_num = int(cell_str)
                existing_screens.append((col, screen_num))
                max_screen_num = max(max_screen_num, screen_num)
            except (ValueError, TypeError):
                pass
            continue
        
        # Check if it starts with an integer
        match = starts_with_int_pattern.match(cell_str)
        if match:
            try:
                screen_num = int(match.group(1))
                existing_screens.append((col, screen_num))
                max_screen_num = max(max_screen_num, screen_num)
            except (ValueError, TypeError):
                pass
    
    existing_count = len(existing_screens)
    
    if existing_count >= required_screens:
        # Already have enough screens
        return
    
    # Need to add more screen columns
    screens_to_add = required_screens - existing_count
    
    # Determine where to insert new headers
    if notes_boundary_col is not None:
        # Insert columns before notes boundary
        insert_col = notes_boundary_col
        ws.insert_cols(insert_col, amount=screens_to_add)
        # Write new headers starting at insert_col
        for i in range(screens_to_add):
            screen_num = max_screen_num + 1 + i
            header_cell = ws.cell(row=header_row, column=insert_col + i, value=screen_num)
            header_cell.fill = HEADER_BLUE
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = GRID_BORDER
    else:
        # No notes boundary - append at end
        # Find first empty column after rightmost_col
        start_col = rightmost_col + 1
        
        # Skip over any occupied columns (never overwrite existing headers)
        while start_col <= ws.max_column:
            cell_value = ws.cell(row=header_row, column=start_col).value
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
                # Found an empty column, use it
                break
            # Column is occupied, skip it
            start_col += 1
        
        # Write new headers
        for i in range(screens_to_add):
            screen_num = max_screen_num + 1 + i
            # Double-check this column is still empty before writing
            check_value = ws.cell(row=header_row, column=start_col).value
            if check_value is not None and (not isinstance(check_value, str) or check_value.strip() != ""):
                # Column is occupied, skip to next
                start_col += 1
                continue
            
            header_cell = ws.cell(row=header_row, column=start_col, value=screen_num)
            header_cell.fill = HEADER_BLUE
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = GRID_BORDER
            start_col += 1
    
    print(f"[INFO] Expanded screen headers from {existing_count} to {required_screens} columns.")


def estimate_required_screens_for_week(
    bookings_df: pd.DataFrame,
    circuit: str,
    theater_name: str,
    city: str,
    state: str,
    curr_week: pd.Timestamp,
) -> int:
    """
    Estimate how many screen columns are needed for a theater's *schedule week*.

    IMPORTANT: The export keeps holdovers on their original Playwk, so the schedule
    for curr_week must include:
      - NEW rows with Playwk == curr_week
      - HOLD rows with Playwk < curr_week (continuations)

    We also exclude:
      - Event-ish rows (units==0 and no FSS)
      - Final rows (ended)
      - "Hold * shows" rows from auto-scheduling (these are treated as special/uncertain)

    Returns at least 1.
    """
    theater_df = bookings_df[
        (bookings_df["Circuit"] == circuit)
        & (bookings_df["Theatre Name"] == theater_name)
        & (bookings_df["City"] == city)
        & (bookings_df["ST"] == state)
    ].copy()

    if theater_df.empty:
        return 1

    theater_df = theater_df[~theater_df.apply(is_event_row, axis=1)].copy()
    if theater_df.empty:
        return 1

    status_norm = theater_df[COL_STATUS].fillna("").astype(str).str.lower()
    is_final = status_norm.str.contains("final")
    is_new = status_norm.str.startswith("new")
    is_hold = status_norm.str.contains("hold")
    is_uncertain_hold = status_norm.str.contains("hold") & status_norm.str.contains(r"\*")

    # Schedule-week inclusion logic
    include_mask = (
        (theater_df[COL_PLAYWK] == curr_week) & is_new
    ) | (
        (theater_df[COL_PLAYWK] < curr_week) & is_hold
    )

    theater_df_main = theater_df[include_mask & (~is_final)].copy()
    if theater_df_main.empty:
        return 1

    theater_df_main["__units_total"] = theater_df_main.apply(compute_units, axis=1)
    theater_df_main = theater_df_main[theater_df_main["__units_total"] > 1e-6].copy()
    if theater_df_main.empty:
        return 1


    layout = build_screen_layout_standard_then_atmos(theater_df_main)
    # build_screen_layout_standard_then_atmos may return (layout, info) when asked; here we only need layout
    if isinstance(layout, tuple):
        layout = layout[0]
    return max(1, len(layout))
def detect_screen_columns(ws: Worksheet, header_row: int = 2) -> Tuple[List[int], Dict[str, int], Optional[int], Optional[int]]:
    """
    Detect screen columns and special columns from a sheet header row.
    
    Scans robustly from column 2 through the last meaningful column, without stopping
    early on blank cells. Detects:
    - Pure integers: 1, 2, 3, ...
    - Headers starting with integer: "10 Atmos", "7 IMAX", "12 RPX", etc.
    - Special columns: "Alternative Content", "Unplayed" (anywhere, not just at end)
    
    Stops only at notes boundary (first column containing "/") or at the rightmost
    non-empty header cell.
    
    Args:
        ws: The worksheet to analyze
        header_row: Row number containing headers (default 2)
    
    Returns:
        Tuple of:
        - List of screen column indices (1-based), in left-to-right order
        - Dict mapping special column names to column indices (1-based)
        - Column index for "Alternative Content" (or None)
        - Column index for "Unplayed" (or None)
    """
    screen_cols: List[int] = []
    special_cols: Dict[str, int] = {}
    alt_content_col: Optional[int] = None
    unplayed_col: Optional[int] = None
    
    if header_row > ws.max_row:
        return screen_cols, special_cols, alt_content_col, unplayed_col
    
    # Regex patterns for screen headers
    pure_int_pattern = re.compile(r'^\s*\d+\s*$')  # Pure integer: "1", "2", "10", etc.
    starts_with_int_pattern = re.compile(r'^\s*\d+\b')  # Starts with integer: "10 Atmos", "7 IMAX", etc.
    
    # First, find the notes boundary (first column containing "/")
    notes_boundary_col: Optional[int] = None
    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is not None:
            cell_str = str(cell_value).strip()
            if "/" in cell_str:
                notes_boundary_col = col
                break
    
    # Find the rightmost non-empty header cell in row 2
    rightmost_col = 2
    for col in range(2, ws.max_column + 1):
        if notes_boundary_col is not None and col >= notes_boundary_col:
            break
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value is not None and (not isinstance(cell_value, str) or cell_value.strip() != ""):
            rightmost_col = col
    
    # Determine scan limit: stop before notes boundary, or at rightmost non-empty header
    scan_limit = notes_boundary_col if notes_boundary_col is not None else (rightmost_col + 1)
    
    # Scan from column 2 up to scan_limit
    col = 2
    while col < scan_limit:
        cell_value = ws.cell(row=header_row, column=col).value
        
        # Skip empty cells (don't stop early)
        if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
            col += 1
            continue
        
        cell_str = str(cell_value).strip()
        
        # Check if it's a pure integer (standard screen)
        if pure_int_pattern.match(cell_str):
            screen_cols.append(col)
            col += 1
            continue
        
        # Check if it starts with an integer (screen with format label)
        if starts_with_int_pattern.match(cell_str):
            screen_cols.append(col)
            col += 1
            continue
        
        # Not a screen header - check if it's a special column
        cell_lower = cell_str.lower()
        if "alternative content" in cell_lower:
            alt_content_col = col
            special_cols["Alternative Content"] = col
            col += 1
            continue
        elif "unplayed" in cell_lower:
            unplayed_col = col
            special_cols["Unplayed"] = col
            col += 1
            continue
        elif cell_str:
            # Other special column - record it but continue scanning
            special_cols[cell_str] = col
            col += 1
            continue
        
        # Unknown header type - skip and continue (don't stop)
        col += 1
    
    return screen_cols, special_cols, alt_content_col, unplayed_col


def process_theater(
    df: pd.DataFrame,
    circuit: str,
    theater_name: str,
    city: str,
    state: str,
    ws: Worksheet,
    screen_cols: List[int],
    unplayed_col: Optional[int],
    alt_content_col: Optional[int],
) -> None:
    """
    Process a single theater's scheduling update.
    
    Separates regular movies from events and updates the sheet accordingly.
    """
    # Get all rows for this theater (including events)
    theater_df = df[
        (df["Circuit"] == circuit)
        & (df["Theatre Name"] == theater_name)
        & (df["City"] == city)
        & (df["ST"] == state)
    ].copy()

    if theater_df.empty:
        print(f"[WARN] No booking rows found for {circuit} / {theater_name}, {city}, {state}.")
        return

    curr_week = theater_df[COL_PLAYWK].max()
    if pd.isna(curr_week):
        print(f"[WARN] Current week missing for {circuit} / {theater_name}, {city}, {state}.")
        return

    prev_candidates = theater_df.loc[theater_df[COL_PLAYWK] < curr_week, COL_PLAYWK]
    prev_week = prev_candidates.max() if not prev_candidates.empty else None

    curr_rows = theater_df[theater_df[COL_PLAYWK] == curr_week]
    prev_rows = (
        theater_df[theater_df[COL_PLAYWK] == prev_week] if prev_week is not None else pd.DataFrame()
    )

    # Separate events from regular rows (row-level detection is more reliable
    # than Status-only rules in these exports).
    if not prev_rows.empty:
        prev_events = prev_rows[prev_rows.apply(is_event_row, axis=1)]
        prev_movies = prev_rows[~prev_rows.apply(is_event_row, axis=1)]
    else:
        prev_events = pd.DataFrame()
        prev_movies = pd.DataFrame()

    if not curr_rows.empty:
        curr_events = curr_rows[curr_rows.apply(is_event_row, axis=1)]
        curr_movies = curr_rows[~curr_rows.apply(is_event_row, axis=1)]
    else:
        curr_events = pd.DataFrame()
        curr_movies = pd.DataFrame()

    # Previous week: Hold + Final movies (each should show up in the prev-week row)
    if not prev_movies.empty:
        hold_mask = prev_movies[COL_STATUS].str.contains("hold", case=False, na=False)
        final_mask = prev_movies[COL_STATUS].str.contains("final", case=False, na=False)
        prev_layout_rows = pd.concat(
            [prev_movies[hold_mask], prev_movies[final_mask]],
            ignore_index=True,
        ).drop_duplicates()
    else:
        hold_mask = pd.Series(dtype=bool)
        prev_layout_rows = pd.DataFrame()
    # Build FSS map used to update the previous-week FSS row (titles layout is preserved).
    # Use ALL rows for this theater (across playwks) because holdovers keep their original Playwk,
    # but the FSS values in the export correspond to the weekend being updated.
    prev_fss_map: dict[str, str] = {}
    if not theater_df.empty:
        non_event = theater_df[~theater_df.apply(is_event_row, axis=1)].copy()
        for _, r in non_event.iterrows():
            t = r.get(COL_TITLE)
            if pd.isna(t):
                continue
            key = normalize_title_key(str(t))
            is_final = "final" in normalize_status(r.get(COL_STATUS))
            val = format_fss_value(r.get(COL_FSS), is_final=is_final)
            if val:
                prev_fss_map[key] = val
    # Current schedule-week layout:
    # The export keeps holdovers on their original Playwk, so for the schedule week (curr_week)
    # we include:
    #   - NEW rows where Playwk == curr_week
    #   - HOLD rows where Playwk < curr_week (continuations)
    # We DO NOT auto-schedule rows like "Hold * shows" (uncertain holds), and we exclude Finals.

    all_movies = theater_df[~theater_df.apply(is_event_row, axis=1)].copy()

    status_norm = all_movies[COL_STATUS].fillna("").astype(str).str.lower()
    is_final = status_norm.str.contains("final")
    is_new = status_norm.str.startswith("new")
    is_hold = status_norm.str.contains("hold")
    is_uncertain_hold = is_hold & status_norm.str.contains(r"\*")

    # Main rows that actually drive screen counts / packing
    include_mask = (
        (all_movies[COL_PLAYWK] == curr_week) & is_new
    ) | (
        (all_movies[COL_PLAYWK] < curr_week) & is_hold
    )

    curr_layout_rows = all_movies[include_mask & (~is_final)].copy()

    # "Hold * shows" rows (often 0.25/0.5 units) are informational add-ons.
    # We DO NOT let them create new screens. Instead we attach ONLY mutually-linked ones
    # (e.g., Truth <-> After) into an existing STANDARD screen based on Comments matching.
    # Adjust WK# for carried holdovers so the display shows the correct run week (wk2, wk3, ...)
    # without changing NEW behavior.
    def bump_wknum(row: pd.Series) -> object:
        wk = row.get(COL_WKNUM)
        pw = row.get(COL_PLAYWK)
        if pd.isna(wk) or pd.isna(pw):
            return wk
        try:
            delta_weeks = int(round((pd.Timestamp(curr_week) - pd.Timestamp(pw)).days / 7))
        except Exception:
            return wk
        if delta_weeks <= 0:
            return wk
        try:
            return int(wk) + delta_weeks
        except (TypeError, ValueError):
            try:
                return float(wk) + float(delta_weeks)
            except Exception:
                return wk

    if not curr_layout_rows.empty:
        curr_layout_rows[COL_WKNUM] = curr_layout_rows.apply(bump_wknum, axis=1)

        # Only rows with actual screen units contribute to screens
        curr_layout_rows["__units_total"] = curr_layout_rows.apply(compute_units, axis=1)
        curr_layout_rows = curr_layout_rows[curr_layout_rows["__units_total"] > 1e-6].copy()


    # Previous week layout: one movie per "screen" column
    prev_layout = [
        [row] for _, row in prev_layout_rows.iterrows()
    ]

    # Current week layout: Standard bucket first, then ATMOS appended at the end
    curr_layout_result = build_screen_layout_standard_then_atmos(curr_layout_rows, return_block_info=True)

    if isinstance(curr_layout_result, tuple):
        curr_layout, block_info = curr_layout_result
    else:
        curr_layout, block_info = curr_layout_result, {"standard_screens": len(curr_layout_result), "premium_blocks": []}

    # Label premium sections (row 1) without changing numeric screen headers
    try:
        apply_premium_section_labels_row1(ws, int(block_info.get("standard_screens", 0)), block_info.get("premium_blocks", []))
    except Exception:
        pass

    # Build per-screen fill map for new movies (first occurrence = yellow, repeats = orange)
    def norm_title(s: str) -> str:
        return (s or "").strip().lower()

    new_seen: Dict[str, int] = {}
    screen_fill_by_idx: Dict[int, PatternFill] = {}

    for screen_idx, screen_rows in enumerate(curr_layout):
        if screen_idx >= len(screen_cols):
            break  # only color actual screen columns
        new_titles = []
        for row in screen_rows:
            status = normalize_status(row.get(COL_STATUS))
            if status.startswith("new"):
                new_titles.append(norm_title(str(row.get(COL_TITLE, ""))))
        if not new_titles:
            continue

        any_first = False
        for t in new_titles:
            count = new_seen.get(t, 0)
            if count == 0:
                any_first = True
            new_seen[t] = count + 1

        screen_fill_by_idx[screen_idx] = NEW_YELLOW if any_first else NEW_ORANGE

# Step 4: Attach split/special titles (units == 0) to a host screen
    # Define split_rows as: units == 0, FSS > 0 (or not null), WK# <= 100 (already filtered >100 in load_bookings)
    split_rows = pd.DataFrame()
    if not curr_movies.empty:
        units_mask = (curr_movies.apply(compute_units, axis=1) == 0)
        # Check for FSS>0 or FSS not null (to catch split movies like "Him final")
        fss_mask = (
            (curr_movies[COL_FSS].notna() & (curr_movies[COL_FSS] > 0)) |
            (curr_movies[COL_STATUS].str.contains("final", case=False, na=False))
        )
        split_mask = units_mask & fss_mask
        
        # WK# <= 100 is already ensured by load_bookings filtering (>100 removed), but double-check
        if COL_WKNUM in curr_movies.columns:
            def is_valid_wk(val):
                if pd.isna(val):
                    return False
                try:
                    wk_int = int(val)
                    return wk_int <= 100
                except (TypeError, ValueError):
                    return False
            wk_mask = curr_movies[COL_WKNUM].apply(is_valid_wk)
            split_mask = split_mask & wk_mask
        
        split_rows = curr_movies[split_mask].copy()
        # Exclude split_rows that are already in curr_layout_rows (to avoid duplicates)
        if not split_rows.empty and not curr_layout_rows.empty:
            split_titles = set(split_rows[COL_TITLE].str.lower().str.strip())
            layout_titles = set(curr_layout_rows[COL_TITLE].str.lower().str.strip())
            split_rows = split_rows[~split_rows[COL_TITLE].str.lower().str.strip().isin(layout_titles)]
    
        # Place each split_row by stacking it into a host screen cell (prepend above host)
    for _, split_row in split_rows.iterrows():
        if not curr_layout and screen_cols:
            # Ensure we have at least one screen container to place split rows
            curr_layout = [[]]

        host_screen_idx = find_host_screen_for_split(split_row, curr_layout)
        if host_screen_idx is not None and host_screen_idx < len(curr_layout):
            curr_layout[host_screen_idx].insert(0, split_row)
        elif curr_layout:
            # Fallback: use first screen
            curr_layout[0].insert(0, split_row)

# Update the sheet with movie layouts
    if prev_layout or curr_layout:
        update_theater_sheet(
            ws, prev_layout, curr_layout, prev_week, curr_week,
            screen_cols, unplayed_col, prev_fss_map,
            alt_content_col=alt_content_col,
            screen_fill_by_idx=screen_fill_by_idx
        )
    
    # Handle Alternative Content (events) if column exists
    if alt_content_col is not None:
        prev_dt = to_excel_datetime(prev_week) if prev_week is not None else None
        curr_dt = to_excel_datetime(curr_week)
        
        # Write events for previous week
        if prev_dt is not None and not prev_events.empty:
            titles_row, _ = ensure_week_rows(ws, prev_dt)
            event_titles = [format_title_line(row) for _, row in prev_events.iterrows()]
            if event_titles:
                # Stack event titles with newlines (use \n\n for readability like regular titles)
                new_events = "\n\n".join(event_titles)
                new_events = sanitize_excel_text(new_events)
                ws.cell(row=titles_row, column=alt_content_col, value=new_events or None)
        
        # Write events for current week
        if not curr_events.empty:
            titles_row, _ = ensure_week_rows(ws, curr_dt)
            event_titles = [format_title_line(row) for _, row in curr_events.iterrows()]
            if event_titles:
                # Stack event titles with newlines (use \n\n for readability like regular titles)
                new_events = "\n\n".join(event_titles)
                new_events = sanitize_excel_text(new_events)
                ws.cell(row=titles_row, column=alt_content_col, value=new_events or None)

    print(f"[INFO] Updated {circuit} / {theater_name}, {city}, {state} ({ws.title}).")


def create_blank_theater_sheet(
    wb: Workbook, 
    sheet_name: str, 
    theater_id: Optional[str] = None,
    default_screens: int = 1
) -> Worksheet:
    """
    Create a blank theater sheet with only headers (no week data).
    
    Row 1: Sheet name as title (bold)
    Row 2: Default numeric screen headers (1, 2, 3, ...)
    AA1: Theater ID metadata (hidden)
    
    Args:
        wb: Workbook to add sheet to
        sheet_name: Name of the theater sheet
        theater_id: Theater ID to store in META_CELL (if provided)
        default_screens: Default number of screen columns (default: 7)
    """
    # Make sheet name safe for Excel
    existing_titles = set(wb.sheetnames)
    safe_sheet_name = make_safe_sheet_title(sheet_name, existing_titles)
    ws = wb.create_sheet(safe_sheet_name)
    
    # Row 1: Title (yellow fill, bold) - use original sheet_name for display
    title_value = sanitize_excel_text(sheet_name)
    title_cell = ws.cell(row=1, column=1, value=title_value)
    title_cell.font = Font(bold=True)
    title_cell.fill = CITY_YELLOW
    
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Row 2: Default headers
    # Column A is blank (date column)
    ws.cell(row=2, column=1, value=None)
    
    # Default headers: use default_screens parameter
    for screen_num in range(1, default_screens + 1):
        header_cell = ws.cell(row=2, column=screen_num + 1, value=screen_num)
        header_cell.fill = HEADER_BLUE
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = GRID_BORDER
    print(f"[INFO] Created sheet '{sheet_name}' with default headers ({default_screens} screens)")
    
    # Write theater_id to META_CELL if provided
    if theater_id:
        ws[META_CELL] = theater_id
        # Hide the metadata column
        ws.column_dimensions[META_COL].hidden = True
    
    return ws



# ----------------------------
# Section labels (premium blocks)
# ----------------------------

SECTION_LABEL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_LABEL_FONT = Font(bold=True)
SECTION_LABEL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

def clear_section_labels_row1(ws: Worksheet, start_col: int = 2) -> None:
    """Remove any merged cells and text labels we previously wrote in row 1 (except A1)."""
    # Unmerge any row-1 merges that start at column >= start_col
    to_unmerge = []
    for r in list(ws.merged_cells.ranges):
        # r is a CellRange
        if r.min_row == 1 and r.max_row == 1 and r.min_col >= start_col:
            to_unmerge.append(str(r))
    for rng in to_unmerge:
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass

    # Clear row 1 cells (except A1)
    max_col = ws.max_column
    for c in range(start_col, max_col + 1):
        cell = ws.cell(row=1, column=c)
        # Only clear if it looks like our label (avoid nuking user-added notes)
        if cell.value is not None:
            cell.value = None
        # reset minimal styling
        cell.fill = PatternFill()
        cell.font = Font()
        cell.alignment = Alignment()
        cell.border = Border()

def apply_premium_section_labels_row1(
    ws: Worksheet,
    standard_screens: int,
    premium_blocks: list[tuple[str, int]],
    header_row: int = 2,
) -> None:
    """
    Write labels for each non-standard (premium) section in row 1 *above* the numbered screen headers.
    Keeps the numeric headers unchanged (so you never get '10 ATMOS' type headers).

    standard_screens: how many screen columns are used by the standard pool (NEW + non-NEW + extras)
    premium_blocks: list of (format_name, screen_count) in the order they are appended
    """
    if standard_screens < 0:
        standard_screens = 0

    # Ensure row 1 exists/styled on A1
    ws.cell(row=1, column=1).fill = CITY_YELLOW
    ws.cell(row=1, column=1).font = Font(bold=True)

    clear_section_labels_row1(ws, start_col=2)

    # Screen columns start at col 2
    col_cursor = 2 + int(standard_screens)

    for fmt, count in premium_blocks:
        if not fmt or count <= 0:
            continue

        start_col = col_cursor
        end_col = col_cursor + int(count) - 1

        # Only label if these columns exist in the sheet
        if start_col > ws.max_column:
            break
        end_col = min(end_col, ws.max_column)

        # Merge label across the block
        if end_col > start_col:
            try:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            except Exception:
                pass

        label_cell = ws.cell(row=1, column=start_col)
        label_cell.value = sanitize_excel_text(str(fmt))
        label_cell.fill = SECTION_LABEL_FILL
        label_cell.font = SECTION_LABEL_FONT
        label_cell.alignment = SECTION_LABEL_ALIGN
        label_cell.border = GRID_BORDER

        # Apply same styling to the full merged span (borders/fill) so it looks consistent
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = SECTION_LABEL_FILL
            cell.font = SECTION_LABEL_FONT
            cell.alignment = SECTION_LABEL_ALIGN
            cell.border = GRID_BORDER

        col_cursor = end_col + 1



def get_or_create_theater_sheet(
    wb: Workbook,
    theater_id: str,
    base_name: str,
    circuit: Optional[str] = None,
    theater_name: Optional[str] = None,
    city: Optional[str] = None,
    state: Optional[str] = None,
    prefix_regex: Optional[re.Pattern] = None,
    include_city: bool = False,
) -> Worksheet:
    """
    Get an existing theater sheet by theater_id, or create a new blank one with headers only.
    
    Uses theater_id stored in META_CELL for stable identification across runs.
    Never relies on sheet title for identity - only for display.
    
    Args:
        wb: Workbook to get/create sheet in
        theater_id: Stable theater ID (from compute_theater_id)
        base_name: Base name of the theater sheet (e.g., "Rotunda, MD") for display
        circuit: Circuit name (for metadata)
        theater_name: Original theater name (for display)
        city: City name (for display)
        state: State code (for display)
        prefix_regex: Optional regex to strip prefixes from theater name
        include_city: Whether to include city in sheet names
    
    Returns:
        Worksheet object (always has META_CELL set with theater_id)
    """
    # FIRST: Look up by theater_id (metadata-based, not title-based)
    existing_sheet = find_sheet_by_theater_id(wb, theater_id)
    if existing_sheet is not None:
        # Found existing sheet - ensure A1 is styled and META_CELL is set
        a1_cell = existing_sheet.cell(row=1, column=1)
        a1_cell.fill = CITY_YELLOW
        a1_cell.font = Font(bold=True)
        # Ensure META_CELL is set (in case it was missing)
        if not existing_sheet[META_CELL].value:
            existing_sheet[META_CELL] = theater_id
            existing_sheet.column_dimensions[META_COL].hidden = True
        return existing_sheet
    
    # NOT FOUND BY ID: Try to find by normalized sheet name key
    # This handles cases where the same theater has different sheet names due to
    # formatting differences (e.g., "Frederick, MD" vs "Frederick, Frederick, MD")
    base_key = normalize_sheet_key(base_name)
    matching_sheets = []
    for sheet_name in wb.sheetnames:
        sheet_key = normalize_sheet_key(sheet_name)
        if sheet_key == base_key:
            matching_sheets.append(sheet_name)
    
    # Reuse the shortest matching sheet name if found
    if matching_sheets:
        reuse_sheet_name = min(matching_sheets, key=len)
        ws = wb[reuse_sheet_name]
        # Update metadata to correct theater_id
        ws[META_CELL] = theater_id
        ws.column_dimensions[META_COL].hidden = True
        # Update A1 display name to match base_name
        a1_cell = ws.cell(row=1, column=1)
        a1_cell.value = sanitize_excel_text(base_name)
        a1_cell.fill = CITY_YELLOW
        a1_cell.font = Font(bold=True)
        return ws
    
    # NOT FOUND: Create new sheet with safe title
    existing_titles = set(wb.sheetnames)
    safe_base_name = make_safe_sheet_title(base_name, existing_titles)
    
    # Create new sheet with theater_id in META_CELL
    min_screens = get_min_screen_count_from_template(theater_name, state)
    ws = create_blank_theater_sheet(wb, safe_base_name, theater_id=theater_id, default_screens=min_screens)
    
    # Ensure META_CELL is set and column is hidden
    ws[META_CELL] = theater_id
    ws.column_dimensions[META_COL].hidden = True
    
    return ws


def check_file_locked(file_path: str) -> bool:
    """
    Check if a file is locked (open in Excel or another process).
    
    Returns True if file is locked, False otherwise.
    """
    if not os.path.exists(file_path):
        return False
    
    try:
        # Try to open the file in append mode to check if it's locked
        with open(file_path, 'a'):
            pass
        return False
    except (PermissionError, IOError):
        return True


def create_blank_master_workbook(
    master_path: str,
    bookings_df: pd.DataFrame,
    prefix_regex: Optional[re.Pattern] = None,
    sheet_name_format: str = "{theatre}, {state}",
    include_city: bool = False,
) -> Workbook:
    """
    Create a new blank master workbook with only headers (no week data).
    
    Creates theater sheets with headers only (rows 1-2), no week data.
    
    Args:
        master_path: Path where the new master file should be created
        bookings_df: DataFrame to determine which theater sheets to create
        prefix_regex: Optional regex to strip prefixes from theater names
        sheet_name_format: Format string for sheet names
        include_city: Whether to include city in sheet names
    
    Returns:
        The created Workbook object
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create theater sheets with headers only
    theaters = get_theaters(bookings_df)
    for _, theater_row in theaters.iterrows():
        circuit = str(theater_row["Circuit"])
        theater_name = str(theater_row["Theatre Name"])
        city = str(theater_row["City"])
        state = str(theater_row["ST"])
        theater_id = compute_theater_id(circuit, theater_name, city, state)
        sheet_name = map_theater_to_sheet_name(
            theater_name, state, city, prefix_regex, sheet_name_format, include_city
        )
        get_or_create_theater_sheet(
            wb, theater_id, sheet_name, circuit, theater_name, city, state, prefix_regex, include_city
        )
    
    # Save the blank workbook
    wb.save(master_path)
    print(f"[INFO] Created blank master workbook '{master_path}' with headers only.")
    
    # Reload to ensure it's properly saved
    wb = load_workbook(master_path)
    return wb


def run_scheduler_update(
    bookings_path: str,
    master_path: str,
    bookings_sheet: str = BOOKINGS_SHEET_NAME,
    prefix_regex: Optional[re.Pattern] = None,
    sheet_name_format: str = "{theatre}, {state}",
    include_city: bool = False,
) -> None:
    """
    Main orchestration routine for updating the master schedule workbook.
    
    Args:
        bookings_path: Path to bookings export Excel file
        master_path: Path to master schedule workbook
        bookings_sheet: Sheet name in bookings workbook
        prefix_regex: Optional regex to strip prefixes from theater names
        sheet_name_format: Format string for sheet names
        include_city: Whether to include city in sheet names
    """
    # Check if output file is locked
    if os.path.exists(master_path) and check_file_locked(master_path):
        print("[ERROR] The master workbook is open in Excel or locked by another process.")
        print("[ERROR] Close the Excel file and try again.")
        return
    
    # Load and clean bookings
    bookings_df = load_bookings(bookings_path, sheet_name=bookings_sheet)
    print(f"[INFO] Loaded {len(bookings_df)} booking rows from '{bookings_path}'.")
    
    # Auto-detect screen unit columns from the bookings DataFrame
    global SCREEN_UNIT_COLS
    SCREEN_UNIT_COLS = detect_screen_unit_cols(bookings_df)
    print(f"[INFO] Detected screen unit columns: {SCREEN_UNIT_COLS}")

    # Establish "standard" (first unit col) and premium cols (remaining), in file order
    global STANDARD_UNIT_COL, PREMIUM_UNIT_COLS
    STANDARD_UNIT_COL = SCREEN_UNIT_COLS[0] if SCREEN_UNIT_COLS else ""
    PREMIUM_UNIT_COLS = SCREEN_UNIT_COLS[1:] if len(SCREEN_UNIT_COLS) > 1 else []

    # Coerce detected unit columns to numeric now that we know what they are
    for col in SCREEN_UNIT_COLS:
        if col in bookings_df.columns:
            bookings_df[col] = pd.to_numeric(bookings_df[col], errors="coerce")
# Load or create master workbook
    if os.path.exists(master_path):
        try:
            wb = load_workbook(master_path)
            print(f"[INFO] Loaded existing master workbook '{master_path}'.")
            
            # Migration: Add theater_id to sheets that don't have it (one-time)
            migrated_count = 0
            for ws in wb.worksheets:
                try:
                    meta_value = ws[META_CELL].value
                    if not meta_value or str(meta_value).strip() == "":
                        # Try to infer theater_id from A1 or sheet title
                        a1_value = ws.cell(row=1, column=1).value
                        inferred_id = None
                        
                        # Try to parse from A1 (e.g., "Rotunda, MD" or "Rotunda (Hagerstown), MD")
                        if a1_value and isinstance(a1_value, str):
                            # Pattern: "{theatre}, {state}" or "{theatre} ({city}), {state}"
                            match = re.match(r'^(.+?)(?:\s*\(([^)]+)\))?,\s*([A-Z]{2})$', a1_value.strip())
                            if match:
                                theatre_part = match.group(1).strip()
                                city_part = match.group(2) if match.group(2) else None
                                state_part = match.group(3).strip()
                                
                                # We can't infer circuit from display name, so leave it blank
                                # Format: "|{theatre}|{city}|{state}" (circuit will be empty)
                                if city_part:
                                    inferred_id = f"|{theatre_part}|{city_part}|{state_part}"
                                else:
                                    inferred_id = f"|{theatre_part}||{state_part}"
                        
                        # If we couldn't infer, leave it blank (will be set correctly when theater is processed)
                        if inferred_id:
                            ws[META_CELL] = inferred_id
                            ws.column_dimensions[META_COL].hidden = True
                            migrated_count += 1
                except (AttributeError, KeyError):
                    # Sheet doesn't have META_CELL yet, skip for now
                    pass
            
            if migrated_count > 0:
                print(f"[INFO] Migrated {migrated_count} sheet(s) to use theater_id metadata.")
        except PermissionError:
            print("[ERROR] Cannot open the master workbook. It may be open in Excel.")
            print("[ERROR] Close the Excel file and try again.")
            return
    else:
        # Create blank workbook with headers only
        wb = create_blank_master_workbook(
            master_path, bookings_df, prefix_regex, sheet_name_format, include_city
        )
    
    # Get unique theaters from the bookings DataFrame
    theaters = get_theaters(bookings_df)
    print(f"[INFO] Found {len(theaters)} unique theaters.")

    for _, theater_row in theaters.iterrows():
        circuit = str(theater_row["Circuit"])
        theater_name = str(theater_row["Theatre Name"])
        city = str(theater_row["City"])
        state = str(theater_row["ST"])
        
        # Compute stable theater_id
        theater_id = compute_theater_id(circuit, theater_name, city, state)

        # Map to sheet name (for display only)
        sheet_name = map_theater_to_sheet_name(
            theater_name, state, city, prefix_regex, sheet_name_format, include_city
        )
        
        # Get or create theater sheet by theater_id (metadata-based lookup)
        ws = get_or_create_theater_sheet(
            wb, theater_id, sheet_name, circuit, theater_name, city, state, prefix_regex, include_city
        )
        
        # Ensure META_CELL is always set with correct theater_id (updates migrated sheets)
        ws[META_CELL] = theater_id
        ws.column_dimensions[META_COL].hidden = True
        
        # Determine current week for this theater to estimate required screens
        theater_df = bookings_df[
            (bookings_df["Circuit"] == circuit)
            & (bookings_df["Theatre Name"] == theater_name)
            & (bookings_df["City"] == city)
            & (bookings_df["ST"] == state)
        ].copy()
        
        curr_week = theater_df[COL_PLAYWK].max() if not theater_df.empty else None
        
        # Estimate required screens for current week
        if curr_week is not None and not pd.isna(curr_week):
            required_screens = estimate_required_screens_for_week(
                bookings_df, circuit, theater_name, city, state, curr_week
            )
            # Minimum screen headers come from template if available (never a max)
            min_screens = get_min_screen_count_from_template(theater_name, state)
            required_screens = max(required_screens, min_screens)
            # Ensure sheet has enough screen headers
            ensure_screen_headers(ws, required_screens, header_row=2)
        
        # Re-detect screen columns after expansion
        screen_cols, special_cols, alt_content_col, unplayed_col = detect_screen_columns(ws, header_row=2)
        # (Premium section labels are written in row 1 during process_theater)
        # Re-detect screen columns (headers are numeric and unchanged)
        screen_cols, special_cols, alt_content_col, unplayed_col = detect_screen_columns(ws, header_row=2)
        
        if not screen_cols:
            print(f"[WARN] No screen columns detected for {sheet_name}. Skipping.")
            continue
        
        print(f"[INFO] Detected {len(screen_cols)} screen columns for {sheet_name}.")
        
        # Style headers (A1 and row 2) with fills, fonts, and borders
        style_headers(ws, screen_cols, alt_content_col, unplayed_col)
        if alt_content_col:
            print(f"[INFO] Alternative Content column found at column {alt_content_col}.")
        if unplayed_col:
            print(f"[INFO] Unplayed column found at column {unplayed_col}.")
        
        # Process theater
        process_theater(
            bookings_df, circuit, theater_name, city, state, ws,
            screen_cols, unplayed_col, alt_content_col
        )
        
        # Re-style FSS cells with Final formatting across all old weeks
        cols_to_style = list(screen_cols)
        if alt_content_col is not None:
            cols_to_style.append(alt_content_col)
        if unplayed_col is not None:
            cols_to_style.append(unplayed_col)
        restyle_final_fss_cells(ws, cols_to_style)

    # Sanitize entire workbook before saving (removes illegal chars from existing content)
    sanitize_workbook_inplace(wb)
    
    # Save workbook
    wb.save(master_path)
    print(f"[INFO] Saved master workbook to '{master_path}'.")
    print("[INFO] All theaters processed.")


def run_scheduler(bookings_path: str, output_dir: str) -> None:
    """
    Legacy main orchestration routine (per-theater outputs).
    
    Kept for backward compatibility if needed.
    """
    bookings_df = load_bookings(bookings_path)
    
    # Auto-detect screen unit columns from the bookings DataFrame
    global SCREEN_UNIT_COLS
    SCREEN_UNIT_COLS = detect_screen_unit_cols(bookings_df)
    print(f"[INFO] Detected screen unit columns: {SCREEN_UNIT_COLS}")
    
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Get unique theaters from the bookings DataFrame
    theaters = get_theaters(bookings_df)

    for _, theater_row in theaters.iterrows():
        circuit = str(theater_row["Circuit"])
        theater_name = str(theater_row["Theatre Name"])
        city = str(theater_row["City"])
        state = str(theater_row["ST"])

        theater_path = make_theater_filename(output_dir, circuit, theater_name, city, state)

        if os.path.exists(theater_path):
            wb = load_workbook(theater_path)
            print(f"[INFO] Loaded existing workbook '{theater_path}'.")
            existing_titles = set(wb.sheetnames)
            safe_schedule_name = make_safe_sheet_title("Schedule", existing_titles)
            if safe_schedule_name in wb.sheetnames:
                ws = wb[safe_schedule_name]
            else:
                ws = wb.create_sheet(safe_schedule_name)
        else:
            wb = Workbook()
            ws = wb.active
            existing_titles = set(wb.sheetnames)
            safe_schedule_name = make_safe_sheet_title("Schedule", existing_titles)
            ws.title = safe_schedule_name
            print(f"[INFO] Created new workbook '{theater_path}' for {circuit} / {theater_name}.")

        # For legacy mode, use default screen columns (1, 2, 3, ...)
        # Default to 7 screens (columns B through H)
        screen_cols = list(range(2, 9))  # Columns B through H (7 screens)
        process_theater(bookings_df, circuit, theater_name, city, state, ws, screen_cols, None, None)

        wb.save(theater_path)
        print(f"[INFO] Saved schedule to '{theater_path}'.")

    print("[INFO] All theaters processed.")


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Weekly cinema scheduler for master schedule workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python cinema_scheduler.py "Bookings Export.xlsx" "Current Schedule.xlsx"
  python cinema_scheduler.py "Bookings.xlsx" "Schedule.xlsx" --bookings-sheet "Data"
  python cinema_scheduler.py "Bookings.xlsx" "Schedule.xlsx" --theater-prefix-regex "^(warehouse|wc)\\s*"
  python cinema_scheduler.py "Bookings.xlsx" "Schedule.xlsx" --include-city-in-sheet-name
        """.strip()
    )
    parser.add_argument(
        "bookings",
        help="Path to bookings export Excel file (e.g., 'Bookings Export.xlsx').",
    )
    parser.add_argument(
        "master",
        help="Path to master schedule workbook (e.g., 'Current Schedule.xlsx').",
    )
    parser.add_argument(
        "--bookings-sheet",
        default=BOOKINGS_SHEET_NAME,
        help=f"Sheet name in bookings workbook (default: '{BOOKINGS_SHEET_NAME}').",
    )
    parser.add_argument(
        "--theater-prefix-regex",
        default=None,
        help="Optional regex pattern to strip prefixes from theater names (e.g., '^(warehouse|wc)\\s*').",
    )
    parser.add_argument(
        "--sheet-name-format",
        default="{theatre}, {state}",
        help="Format string for sheet names (default: '{theatre}, {state}').",
    )
    parser.add_argument(
        "--include-city-in-sheet-name",
        action="store_true",
        help="Include city in sheet names to avoid collisions (format becomes '{theatre} ({city}), {state}').",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="After processing, scan workbook for any remaining illegal characters and report them (does not modify workbook).",
    )
    return parser.parse_args(argv)


def main() -> None:
    """Entry point."""
    args = parse_args()
    
    # Compile prefix regex if provided
    prefix_regex = None
    if args.theater_prefix_regex:
        try:
            prefix_regex = re.compile(args.theater_prefix_regex, re.IGNORECASE)
        except re.error as e:
            print(f"[ERROR] Invalid regex pattern: {e}")
            return
    
    # Adjust sheet name format if include_city is True
    sheet_name_format = args.sheet_name_format
    if args.include_city_in_sheet_name:
        sheet_name_format = "{theatre} ({city}), {state}"
    
    run_scheduler_update(
        bookings_path=args.bookings,
        master_path=args.master,
        bookings_sheet=args.bookings_sheet,
        prefix_regex=prefix_regex,
        sheet_name_format=sheet_name_format,
        include_city=args.include_city_in_sheet_name,
    )
    
    # Run validation if requested
    if args.validate:
        print("[INFO] Running validation scan...")
        try:
            wb = load_workbook(args.master)
            issues = validate_workbook(wb, max_reports=10)
            if issues:
                print(f"[WARN] Found {len(issues)} cell(s) with illegal characters:")
                for sheet_name, coord, preview in issues:
                    print(f"  - {sheet_name}!{coord}: {repr(preview)}")
            else:
                print("[INFO] No illegal characters found in workbook.")
            wb.close()
        except Exception as e:
            print(f"[ERROR] Validation failed: {e}")


if __name__ == "__main__":
    main()
